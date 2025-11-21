import streamlit as st
import pandas as pd
from streamlit_echarts import st_echarts
from datetime import datetime, timedelta
import json
from pathlib import Path
import time
import os
import logging
import base64
import re
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill
import io
import zipfile
import hashlib

st.set_page_config(
    page_title="Employee Progress Tracker",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# File paths for local storage
USERS_FILE = 'users.json'
ADMINS_FILE = 'admins.json'
ATTENDANCE_FILE = 'attendance.json'

# Helper functions for authentication and data management
def hash_password(password: str) -> str:
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

def load_users():
    """Load users from JSON file"""
    if Path(USERS_FILE).exists():
        try:
            with open(USERS_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_users(users):
    """Save users to JSON file"""
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f, indent=2)

def load_admins():
    """Load admins from JSON file"""
    if Path(ADMINS_FILE).exists():
        try:
            with open(ADMINS_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    # Create default admin if file doesn't exist
    default_admins = {
        "admin": {
            "username": "admin",
            "password": hash_password("admin123"),
            "name": "System Administrator"
        }
    }
    save_admins(default_admins)
    return default_admins

def save_admins(admins):
    """Save admins to JSON file"""
    with open(ADMINS_FILE, 'w') as f:
        json.dump(admins, f, indent=2)

def load_attendance():
    """Load attendance records from JSON file"""
    if Path(ATTENDANCE_FILE).exists():
        try:
            with open(ATTENDANCE_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_attendance(attendance):
    """Save attendance records to JSON file"""
    with open(ATTENDANCE_FILE, 'w') as f:
        json.dump(attendance, f, indent=2)

def add_user(username, password, name, emp_id, email=None, department=None, role=None):
    """Add new user"""
    users = load_users()
    if username in users:
        raise Exception("User already exists")
    
    users[username] = {
        "username": username,
        "password": hash_password(password),
        "name": name,
        "emp_id": emp_id,
        "email": email,
        "department": department,
        "role": role,
        "created_at": datetime.now().isoformat()
    }
    save_users(users)
    return users[username]

def verify_user(username, password):
    """Verify user credentials"""
    users = load_users()
    if username in users:
        if users[username]["password"] == hash_password(password):
            return users[username]
    return None

def verify_admin(username, password):
    """Verify admin credentials"""
    admins = load_admins()
    if username in admins:
        if admins[username]["password"] == hash_password(password):
            return admins[username]
    return None

def checkin_attendance(emp_id, date_str, time_str):
    """Record check-in"""
    attendance = load_attendance()
    key = f"{emp_id}_{date_str}"
    
    if key not in attendance:
        attendance[key] = {
            "emp_id": emp_id,
            "date": date_str,
            "checkin_time": time_str,
            "checkout_time": None
        }
    else:
        attendance[key]["checkin_time"] = time_str
    
    save_attendance(attendance)
    return attendance[key]

def checkout_attendance(emp_id, date_str, time_str):
    """Record check-out"""
    attendance = load_attendance()
    key = f"{emp_id}_{date_str}"
    
    if key in attendance:
        attendance[key]["checkout_time"] = time_str
    else:
        attendance[key] = {
            "emp_id": emp_id,
            "date": date_str,
            "checkin_time": None,
            "checkout_time": time_str
        }
    
    save_attendance(attendance)
    return attendance[key]

def get_attendance_record(emp_id, date_str):
    """Get attendance record for specific date"""
    attendance = load_attendance()
    key = f"{emp_id}_{date_str}"
    return attendance.get(key, None)

def get_user_attendance_history(emp_id, limit=30):
    """Get attendance history for a user"""
    attendance = load_attendance()
    user_records = []
    
    for key, record in attendance.items():
        if record["emp_id"] == emp_id:
            user_records.append(record)
    
    # Sort by date descending
    user_records.sort(key=lambda x: x["date"], reverse=True)
    return user_records[:limit]

def get_all_users():
    """Get all registered users"""
    return load_users()

def get_today_attendance():
    """Get today's attendance records"""
    today_str = datetime.now().date().isoformat()
    attendance = load_attendance()
    today_records = []
    
    for key, record in attendance.items():
        if record["date"] == today_str:
            today_records.append(record)
    
    return today_records
# Helper function to convert image to base64
def get_base64_image(image_path):
    """Convert image to base64 for CSS background"""
    try:
        with open(image_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    except:
        return None
# Custom CSS (dark/black background)
st.markdown("""
<style>
    /* Background styling */
    .stApp {
        background: #000000; /* black */
        background-attachment: fixed;
        color: #e6eef2; /* light default text color for readability */
    }
    /* Subtle pattern overlay (very light) */
    .stApp::before {
        content: "";
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        pointer-events: none;
        z-index: 0;
    }
    /* Main content area */
    .main > div {
        padding: 1rem;
        position: relative;
        z-index: 1;
    }
    /* Block container styling (dark) */
    .block-container {
        padding: 2rem 1rem;
        background: rgba(10, 10, 10, 0.75);
        border-radius: 15px;
        backdrop-filter: blur(6px);
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.6);
        color: #e6eef2; /* ensure text inside blocks is light */
    }
    /* Metric cards */
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px;
        border-radius: 10px;
        color: white !important;
        text-align: center;
        margin: 10px 0;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        transition: transform 0.25s ease;
    }
    .metric-card:hover {
        transform: translateY(-4px);
    }
    .metric-value {
        font-size: 2.5rem;
        font-weight: bold;
    }
    .metric-label {
        font-size: 1rem;
        opacity: 0.95;
    }
    /* Filter container (dark) */
    .filter-container {
        background: linear-gradient(180deg, rgba(20,20,20,0.6) 0%, rgba(30,30,30,0.6) 100%);
        padding: 20px;
        border-radius: 10px;
        margin-bottom: 20px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.6);
        color: #e6eef2;
    }
    /* Button styling */
    .stButton > button {
        width: 100%;
        border-radius: 6px;
        height: 3rem;
        font-weight: 600;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white !important;
        border: none;
        transition: all 0.2s ease;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 18px rgba(102, 126, 234, 0.18);
    }
    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #4c5bd4 0%, #6b4bb8 100%);
    }
    [data-testid="stSidebar"] * {
        color: white !important;
    }
    /* Input fields (dark theme) */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stSelectbox > div > div > select {
        border-radius: 8px;
        border: 1px solid rgba(255,255,255,0.12);
        transition: border-color 0.3s ease;
        background: #0f1113;
        color: #e6eef2;
    }
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus,
    .stSelectbox > div > div > select:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.06);
    }
    /* Logo container styling */
    .logo-container {
        background: transparent;
        padding: 20px;
        border-radius: 15px;
        box-shadow: none;
        margin-bottom: 20px;
        text-align: center;
    }
    /* Constrain logo image size so it doesn't take the entire viewport */
    .logo-container img {
        max-width: 480px;
        width: 100%;
        height: auto;
        display: inline-block;
    }
    /* Expander styling */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, #1e293b 0%, #111827 100%);
        border-radius: 8px;
        font-weight: 600;
        color: #e6eef2 !important;
    }
    @media (max-width: 768px) {
        .main > div {
            padding: 0.5rem;
        }
        .metric-value {
            font-size: 1.8rem;
        }
        .block-container {
            padding: 1rem 0.5rem;
        }
    }
</style>
""", unsafe_allow_html=True)
# Constants
EXCEL_FILE_PATH = r'task_tracker.xlsx'
CONFIG_FILE = 'config.json'
DATA_COLUMNS = [
    'Date',
    'Work Mode',
    'Emp Id',
    'Name',
    'Project Name',
    'Task Title',
    'Task Assigned By',
    'Task Priority',
    'Task Status',
    'Plan for next day',
    'Support Request',
    'Availability',
    'Effort (in hours)',
    'Employee Performance (%)'
]
SUMMARY_SHEET_NAME = '📈 Employee Progress Dashboard'
PERFORMANCE_SHEET_NAME = 'Employee Performance'
WEEKLY_SHEET_NAME = '📊 Weekly Progress Dashboard'
EMPLOYEE_SHEET_SUFFIX = ' Dashboard'
def sanitize_sheet_name(name: str) -> str:
    """Return a workbook-safe base sheet name (<=31 chars, invalid chars removed)."""
    safe = re.sub(r'[\\/*?:\[\]]', '_', str(name)).strip()
    if not safe:
        safe = 'Unnamed'
    return safe[:31]
def build_employee_sheet_name(base_name: str, used_names: set[str]) -> str:
    """Construct a unique sheet name for an employee while respecting Excel limits."""
    suffix = EMPLOYEE_SHEET_SUFFIX
    max_base_len = max(0, 31 - len(suffix))
    trimmed_base = base_name[:max_base_len] if max_base_len else base_name[:31]
    candidate = f"{trimmed_base}{suffix}"
    counter = 2
    while candidate in used_names:
        extra = f" {counter}"
        counter += 1
        allowed_len = max(0, 31 - len(suffix) - len(extra))
        trimmed_base = base_name[:allowed_len] if allowed_len else ''
        fallback = trimmed_base if trimmed_base else 'Employee'
        candidate = f"{fallback}{extra}{suffix}"
    used_names.add(candidate)
    return candidate
def ensure_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Guarantee the performance and effort columns exist and are numeric."""
    df = df.copy()
    numeric_cols = ['Employee Performance (%)', 'Effort (in hours)']
    for col in numeric_cols:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = (
            pd.to_numeric(df[col], errors='coerce')
            .fillna(0.0)
            .astype(float)
        )
    return df
# ==================== PERFORMANCE CALCULATION ====================
def calculate_performance(tasks_list):
    """
    Calculate employee performance using formula:
    Average % = (Sum of Task Priority / Total Effort in Hours) * 100
   
    Priority weights: Low=1, Medium=2, High=3, Critical=4
    """
    if not tasks_list:
        return 0.0
   
    priority_weights = {
        'Low': 1,
        'Medium': 2,
        'High': 3,
        'Critical': 4
    }
   
    total_priority_weight = 0
    total_effort = 0
   
    for task in tasks_list:
        priority = task.get('Task Priority', 'Low')
        try:
            effort = float(task.get('Effort (in hours)', 0))
        except Exception:
            effort = 0.0
       
        weight = priority_weights.get(priority, 1)
        total_priority_weight += weight
        total_effort += effort
   
    if total_effort == 0:
        return 0.0
   
    performance = (total_priority_weight / total_effort) * 100
    return min(round(performance, 2), 100.0) # Cap at 100%
def update_dashboard_sheets(excel_path: str, full_df: pd.DataFrame) -> None:
    """Regenerate the summary and individual employee dashboard sheets."""
    if full_df is None or full_df.empty:
        logging.info("Skipping dashboard sheet update because there is no data.")
        return
    if 'Name' not in full_df.columns:
        logging.warning("Cannot build dashboard sheets because 'Name' column is missing.")
        return
    try:
        full_df = ensure_numeric_columns(full_df)
        if 'Date' in full_df.columns:
            full_df['Date'] = pd.to_datetime(full_df['Date'], errors='coerce')
    except Exception as parse_error:
        logging.error(f"Failed to normalise data for dashboard sheets: {parse_error}")
        return
    try:
        book = load_workbook(excel_path)
    except Exception as workbook_error:
        logging.error(f"Unable to open workbook '{excel_path}' to update dashboard sheets: {workbook_error}")
        return
    # Clean up existing dashboard-related sheets
    all_sheetnames = list(book.sheetnames)
    for sheet_name in all_sheetnames:
        if sheet_name == SUMMARY_SHEET_NAME:
            del book[sheet_name]
        elif sheet_name == PERFORMANCE_SHEET_NAME:
            del book[sheet_name]
        elif sheet_name == WEEKLY_SHEET_NAME:
            del book[sheet_name]
        elif sheet_name.endswith(EMPLOYEE_SHEET_SUFFIX) and sheet_name != SUMMARY_SHEET_NAME:
            del book[sheet_name]
    # Prepare data for summary
    summary_records = []
    unique_names = (
        full_df['Name']
        .dropna()
        .astype(str)
        .str.strip()
    )
    unique_names = [name for name in unique_names.unique() if name]
    for name in unique_names:
        emp_mask = full_df['Name'].astype(str).str.strip() == name
        emp_data = full_df[emp_mask]
        total_tasks = len(emp_data)
        if 'Task Status' in emp_data.columns:
            completed_tasks = int((emp_data['Task Status'] == 'Completed').sum())
        else:
            completed_tasks = 0
        pending_tasks = max(total_tasks - completed_tasks, 0)
        completion_rate = round((completed_tasks / total_tasks * 100) if total_tasks else 0.0, 2)
        avg_perf = round(emp_data['Employee Performance (%)'].mean(), 2)
        last_update = None
        if 'Date' in emp_data.columns and not emp_data['Date'].dropna().empty:
            last_update = emp_data['Date'].dropna().max()
        summary_records.append({
            'name': name,
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'pending_tasks': pending_tasks,
            'completion_rate': completion_rate,
            'avg_performance': avg_perf,
            'last_update': last_update
        })
    # Sort by average performance descending, then by completion rate
    summary_records.sort(key=lambda record: (record['avg_performance'], record['completion_rate']), reverse=True)
    ws_summary = book.create_sheet(SUMMARY_SHEET_NAME)
    summary_headers = [
        'Employee Name',
        'Total Tasks',
        'Completed Tasks',
        'Pending Tasks',
        'Completion Rate (%)',
        'Employee Performance (%)',
        'Last Update',
        'Individual Dashboard'
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        ws_summary.cell(row=1, column=col_idx, value=header)
    ws_summary.freeze_panes = "A2"
    col_widths = [28, 14, 16, 14, 20, 20, 16, 24]
    for idx, width in enumerate(col_widths, start=1):
        column_letter = ws_summary.cell(row=1, column=idx).column_letter
        ws_summary.column_dimensions[column_letter].width = width
    used_sheet_names: set[str] = set(book.sheetnames)
    data_start_row = 2
    for offset, record in enumerate(summary_records):
        row_idx = data_start_row + offset
        ws_summary.cell(row=row_idx, column=1, value=record['name'])
        ws_summary.cell(row=row_idx, column=2, value=record['total_tasks'])
        ws_summary.cell(row=row_idx, column=3, value=record['completed_tasks'])
        ws_summary.cell(row=row_idx, column=4, value=record['pending_tasks'])
        ws_summary.cell(row=row_idx, column=5, value=record['completion_rate'])
        ws_summary.cell(row=row_idx, column=6, value=record['avg_performance'])
        last_update_value = ""
        if record['last_update'] is not None and not pd.isna(record['last_update']):
            if isinstance(record['last_update'], pd.Timestamp):
                last_update_value = record['last_update'].date().isoformat()
            else:
                last_update_value = str(record['last_update'])
        ws_summary.cell(row=row_idx, column=7, value=last_update_value)
        base_name = sanitize_sheet_name(record['name'])
        employee_sheet_name = build_employee_sheet_name(base_name, used_sheet_names)
        hyperlink_formula = f'=HYPERLINK("#\'{employee_sheet_name}\'!A1", "View Dashboard")'
        ws_summary.cell(row=row_idx, column=8).value = hyperlink_formula
        # Build individual employee sheet
        ws_emp = book.create_sheet(employee_sheet_name)
        ws_emp.freeze_panes = "A8"
        ws_emp.cell(row=1, column=1, value=f"Employee Dashboard")
        ws_emp.cell(row=2, column=1, value="Employee Name")
        ws_emp.cell(row=2, column=2, value=record['name'])
        ws_emp.cell(row=3, column=1, value="Total Tasks")
        ws_emp.cell(row=3, column=2, value=record['total_tasks'])
        ws_emp.cell(row=4, column=1, value="Completed Tasks")
        ws_emp.cell(row=4, column=2, value=record['completed_tasks'])
        ws_emp.cell(row=5, column=1, value="Pending Tasks")
        ws_emp.cell(row=5, column=2, value=record['pending_tasks'])
        ws_emp.cell(row=6, column=1, value="Completion Rate (%)")
        ws_emp.cell(row=6, column=2, value=record['completion_rate'])
        ws_emp.cell(row=7, column=1, value="Avg Performance (%)")
        ws_emp.cell(row=7, column=2, value=record['avg_performance'])
        ws_emp.cell(row=2, column=4, value="Last Update")
        ws_emp.cell(row=2, column=5, value=last_update_value)
        ws_emp.cell(row=3, column=4, value="Back to Dashboard")
        ws_emp.cell(row=3, column=5).value = f'=HYPERLINK("#\'{SUMMARY_SHEET_NAME}\'!A1", "View All Employees")'
        ws_emp.cell(row=9, column=1, value="Task Details")
        header_row = 10
        detail_start_row = header_row + 1
        emp_details = full_df[emp_mask].copy()
        emp_details = emp_details.sort_values(by='Date') if 'Date' in emp_details.columns else emp_details
        for col_idx, col_name in enumerate(DATA_COLUMNS, start=1):
            ws_emp.cell(row=header_row, column=col_idx, value=col_name)
        for row_offset, (_, detail_row) in enumerate(emp_details.iterrows()):
            excel_row_idx = detail_start_row + row_offset
            for col_idx, col_name in enumerate(DATA_COLUMNS, start=1):
                cell_value = detail_row.get(col_name)
                if pd.isna(cell_value):
                    cell_value = ""
                elif isinstance(cell_value, pd.Timestamp):
                    cell_value = cell_value.date()
                ws_emp.cell(row=excel_row_idx, column=col_idx, value=cell_value)
        for col_idx in range(1, len(DATA_COLUMNS) + 1):
            column_letter = ws_emp.cell(row=header_row, column=col_idx).column_letter
            ws_emp.column_dimensions[column_letter].width = 18
    if summary_records:
        ws_summary.auto_filter.ref = f"A1:H{data_start_row + len(summary_records) - 1}"
        ws_perf = book.create_sheet(PERFORMANCE_SHEET_NAME)
        perf_headers = [
            'Rank',
            'Employee Name',
            'Total Tasks',
            'Completed Tasks',
            'Completion Rate (%)',
            'Employee Performance (%)',
            'Last Update',
            'Dashboard Link'
        ]
        perf_col_widths = [8, 28, 14, 16, 20, 20, 16, 24]
        for col_idx, header in enumerate(perf_headers, start=1):
            ws_perf.cell(row=1, column=col_idx, value=header)
            column_letter = ws_perf.cell(row=1, column=col_idx).column_letter
            width = perf_col_widths[col_idx - 1] if col_idx - 1 < len(perf_col_widths) else 18
            ws_perf.column_dimensions[column_letter].width = width
        ws_perf.freeze_panes = "A2"
        for rank, record in enumerate(summary_records, start=1):
            row_idx = rank + 1
            ws_perf.cell(row=row_idx, column=1, value=rank)
            ws_perf.cell(row=row_idx, column=2, value=record['name'])
            ws_perf.cell(row=row_idx, column=3, value=record['total_tasks'])
            ws_perf.cell(row=row_idx, column=4, value=record['completed_tasks'])
            ws_perf.cell(row=row_idx, column=5, value=record['completion_rate'])
            ws_perf.cell(row=row_idx, column=6, value=record['avg_performance'])
            last_update_value = ws_summary.cell(data_start_row + rank - 1, column=7).value
            ws_perf.cell(row=row_idx, column=7, value=last_update_value)
            ws_perf.cell(row=row_idx, column=8).value = f'=HYPERLINK("#\'{SUMMARY_SHEET_NAME}\'!A{data_start_row + rank - 1}", "Open Dashboard")'
        ws_perf.auto_filter.ref = f"A1:H{len(summary_records) + 1}"

    # Create Weekly Progress Dashboard
    today = datetime.now().date()
    week_start = today - timedelta(days=6)
    weekly_df = full_df[
        (full_df['Date'].dt.date >= week_start) &
        (full_df['Date'].dt.date <= today)
    ].copy()

    weekly_summary_records = []
    for name in unique_names:
        emp_weekly_mask = weekly_df['Name'].astype(str).str.strip() == name
        emp_weekly = weekly_df[emp_weekly_mask]
        total_tasks_week = len(emp_weekly)
        if 'Task Status' in emp_weekly.columns:
            completed_week = int((emp_weekly['Task Status'] == 'Completed').sum())
        else:
            completed_week = 0
        pending_week = max(total_tasks_week - completed_week, 0)
        completion_rate_week = round((completed_week / total_tasks_week * 100) if total_tasks_week else 0.0, 2)
        avg_perf_week = round(emp_weekly['Employee Performance (%)'].mean(), 2)
        total_effort_week = round(emp_weekly['Effort (in hours)'].sum(), 1)
        workload_status = 'Unknown'
        if 'Availability' in emp_weekly.columns and not emp_weekly.empty:
            avail_counts = emp_weekly['Availability'].value_counts()
            if not avail_counts.empty:
                workload_status = avail_counts.index[0]  # Most common availability
        weekly_summary_records.append({
            'name': name,
            'total_tasks': total_tasks_week,
            'completed_tasks': completed_week,
            'pending_tasks': pending_week,
            'completion_rate': completion_rate_week,
            'avg_performance': avg_perf_week,
            'total_effort': total_effort_week,
            'workload_status': workload_status
        })

    # Sort by average performance descending
    weekly_summary_records.sort(key=lambda record: record['avg_performance'], reverse=True)

    # Overall weekly metrics
    overall_total_tasks = len(weekly_df)
    overall_completed = int((weekly_df['Task Status'] == 'Completed').sum()) if 'Task Status' in weekly_df.columns else 0
    overall_completion = round((overall_completed / overall_total_tasks * 100) if overall_total_tasks else 0.0, 2)
    overall_avg_perf = round(weekly_df['Employee Performance (%)'].mean(), 2)
    overall_total_effort = round(weekly_df['Effort (in hours)'].sum(), 1)

    ws_weekly = book.create_sheet(WEEKLY_SHEET_NAME)
    ws_weekly.freeze_panes = "A7"  # Freeze above the table

    # Title
    ws_weekly.merge_cells('A1:I1')
    ws_weekly.cell(row=1, column=1).value = f"📊 Weekly Progress Dashboard - Week of {week_start.strftime('%Y-%m-%d')} to {today.strftime('%Y-%m-%d')}"
    ws_weekly.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    # Back to Summary Link
    ws_weekly.cell(row=2, column=1, value="Back to Overall Dashboard")
    ws_weekly.cell(row=2, column=2).value = f'=HYPERLINK("#\'{SUMMARY_SHEET_NAME}\'!A1", "View All-Time Summary")'

    # Overall Metrics Section
    ws_weekly.cell(row=3, column=1, value="Overall Weekly Metrics")
    ws_weekly.cell(row=3, column=1).font = openpyxl.styles.Font(bold=True)
    ws_weekly.merge_cells('A3:B3')

    metrics_start_row = 4
    overall_metrics = [
        ('Total Tasks', overall_total_tasks),
        ('Completed Tasks', overall_completed),
        ('Overall Completion Rate (%)', overall_completion),
        ('Average Performance (%)', overall_avg_perf),
        ('Total Effort (Hours)', overall_total_effort)
    ]
    for idx, (label, value) in enumerate(overall_metrics):
        row = metrics_start_row + idx
        ws_weekly.cell(row=row, column=1, value=label)
        ws_weekly.cell(row=row, column=2, value=value)

    # Weekly Table Headers (starting after overall metrics)
    table_start_row = metrics_start_row + len(overall_metrics) + 1
    weekly_headers = [
        'Employee Name',
        'Total Tasks (Week)',
        'Completed',
        'Pending',
        'Completion Rate (%)',
        'Avg Performance (%)',
        'Total Effort (hrs)',
        'Workload Status',
        'Individual Dashboard'
    ]
    for col_idx, header in enumerate(weekly_headers, start=1):
        ws_weekly.cell(row=table_start_row, column=col_idx, value=header)
        ws_weekly.cell(row=table_start_row, column=col_idx).font = openpyxl.styles.Font(bold=True)

    # Column widths for weekly sheet
    weekly_col_widths = [28, 18, 14, 14, 20, 20, 16, 16, 24]
    for idx, width in enumerate(weekly_col_widths, start=1):
        column_letter = ws_weekly.cell(row=table_start_row, column=idx).column_letter
        ws_weekly.column_dimensions[column_letter].width = width

    # Populate weekly table
    for offset, record in enumerate(weekly_summary_records):
        row_idx = table_start_row + 1 + offset
        ws_weekly.cell(row=row_idx, column=1, value=record['name'])
        ws_weekly.cell(row=row_idx, column=2, value=record['total_tasks'])
        ws_weekly.cell(row=row_idx, column=3, value=record['completed_tasks'])
        ws_weekly.cell(row=row_idx, column=4, value=record['pending_tasks'])
        ws_weekly.cell(row=row_idx, column=5, value=record['completion_rate'])
        ws_weekly.cell(row=row_idx, column=6, value=record['avg_performance'])
        ws_weekly.cell(row=row_idx, column=7, value=record['total_effort'])
        ws_weekly.cell(row=row_idx, column=8, value=record['workload_status'])
        base_name = sanitize_sheet_name(record['name'])
        employee_sheet_name = build_employee_sheet_name(base_name, used_sheet_names)
        hyperlink_formula = f'=HYPERLINK("#\'{employee_sheet_name}\'!A1", "View Dashboard")'
        ws_weekly.cell(row=row_idx, column=9).value = hyperlink_formula

    # Auto-filter for weekly table
    if weekly_summary_records:
        last_row = table_start_row + len(weekly_summary_records)
        ws_weekly.auto_filter.ref = f"A{table_start_row}:I{last_row}"

    try:
        book.save(excel_path)
    except Exception as save_error:
        logging.error(f"Failed to save workbook with updated dashboard sheets: {save_error}")
# Helper Functions
def load_config():
    """Load configuration from file"""
    if Path(CONFIG_FILE).exists():
        with open(CONFIG_FILE, 'r') as f:
            return json.load(f)
    return {
        'excel_file_path': EXCEL_FILE_PATH,
        'logo_path': '/home/pinku/PTF Track/logo/PTF1.png',
        'reminder_time': '18:00',
        'reminder_days': [0, 1, 2, 3, 4, 5], # Mon-Sat
        'admin_email': '',
        'employee_emails': []
    }
def save_config(config):
    """Save configuration to file"""
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=4)
def read_excel_data(excel_path=None):
    """Read data from local Excel file"""
    if excel_path is None:
        excel_path = EXCEL_FILE_PATH
   
    try:
        if not os.path.exists(excel_path):
            # Create empty Excel file with headers if it doesn't exist
            df = pd.DataFrame(columns=DATA_COLUMNS)
            df.to_excel(excel_path, index=False, engine='openpyxl')
            return df
       
        # Read Excel file
        df = pd.read_excel(excel_path, engine='openpyxl')
       
        # Handle empty file
        if df.empty:
            return pd.DataFrame()
       
        # Ensure 'Employee Performance (%)' column exists
        return ensure_numeric_columns(df)
   
    except Exception as error:
        st.error(f"Error reading Excel file: {error}")
        return None
def append_to_excel(data_list, excel_path=None):
    """Append data to local Excel file with retry logic for concurrent access
    Args:
        data_list: List of dictionaries, each representing a row to append
    """
    if excel_path is None:
        excel_path = EXCEL_FILE_PATH
   
    max_retries = 3
    retry_delay = 0.5
   
    for attempt in range(max_retries):
        try:
            # Check if file exists and is accessible
            if os.path.exists(excel_path):
                # Try to check if file is locked by attempting to open it
                try:
                    # Try to read the file first to check if it's locked
                    existing_df = pd.read_excel(excel_path, engine='openpyxl')
                    # Remove any completely empty rows
                    existing_df = existing_df.dropna(how='all')
                except PermissionError as pe:
                    if attempt < max_retries - 1:
                        logging.warning(f"Permission error on attempt {attempt + 1}, retrying...")
                        time.sleep(retry_delay * (attempt + 1))
                        continue
                    else:
                        st.error(f"❌ Permission Error: Excel file is locked or inaccessible.")
                        st.error(f"📁 File path: {excel_path}")
                        st.error(f"💡 Please ensure:")
                        st.error(f" 1. The Excel file is not open in Excel or another program")
                        st.error(f" 2. You have write permissions to the file")
                        st.error(f" 3. No other process is using the file")
                        st.error(f" Error details: {str(pe)}")
                        return False
                except Exception as e:
                    # If file is corrupted or empty, start fresh
                    logging.warning(f"Error reading file, starting fresh: {e}")
                    existing_df = pd.DataFrame()
            else:
                existing_df = pd.DataFrame()
           
            # Create new rows DataFrame
            new_rows = pd.DataFrame(data_list)
           
            # Define column order
            columns = DATA_COLUMNS
           
            # Combine with existing data
            if existing_df.empty:
                for col in columns:
                    if col not in new_rows.columns:
                        new_rows[col] = 0.0 if col == 'Employee Performance (%)' else ''
                combined_df = new_rows[columns]
            else:
                # Ensure column order matches
                # Add missing columns if any
                for col in columns:
                    if col not in existing_df.columns:
                        existing_df[col] = 0.0 if col == 'Employee Performance (%)' else ''
                    if col not in new_rows.columns:
                        new_rows[col] = 0.0 if col == 'Employee Performance (%)' else ''
               
                # Reorder columns
                existing_df = existing_df[columns]
                new_rows = new_rows[columns]
               
                combined_df = pd.concat([existing_df, new_rows], ignore_index=True)
           
            # Ensure all columns are in the right order
            if not existing_df.empty:
                combined_df = combined_df[columns]
            combined_df = ensure_numeric_columns(combined_df)
           
            # Write to Excel
            try:
                with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                    combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
            except PermissionError as pe:
                if attempt < max_retries - 1:
                    logging.warning(f"Permission error writing file on attempt {attempt + 1}, retrying...")
                    time.sleep(retry_delay * (attempt + 1))
                    continue
                else:
                    st.error(f"❌ Permission Error: Cannot write to Excel file.")
                    st.error(f"📁 File path: {excel_path}")
                    st.error(f"💡 Please ensure:")
                    st.error(f" 1. The Excel file is not open in Excel or another program")
                    st.error(f" 2. You have write permissions to the file and directory")
                    st.error(f" 3. No other process is using the file")
                    st.error(f" Error details: {str(pe)}")
                    return False
            except Exception as write_error:
                if attempt < max_retries - 1:
                    logging.warning(f"Write error on attempt {attempt + 1}, retrying...")
                    time.sleep(retry_delay * (attempt + 1))
                    continue
                else:
                    st.error(f"❌ Error writing to Excel file: {str(write_error)}")
                    st.error(f"📁 File path: {excel_path}")
                    st.error(f"💡 Error type: {type(write_error).__name__}")
                    return False
            # Update dashboard sheets after successful main sheet write
            try:
                update_dashboard_sheets(excel_path, combined_df)
            except Exception as dash_error:
                logging.error(f"Failed to update dashboard sheets: {dash_error}")
                # Continue without failing the main append
           
            return True
       
        except PermissionError as pe:
            # File might be locked by another process
            if attempt < max_retries - 1:
                logging.warning(f"Permission error on attempt {attempt + 1}, retrying...")
                time.sleep(retry_delay * (attempt + 1))
                continue
            else:
                st.error(f"❌ Permission Error: Excel file is locked or inaccessible.")
                st.error(f"📁 File path: {excel_path}")
                st.error(f"💡 Please ensure:")
                st.error(f" 1. The Excel file is not open in Excel or another program")
                st.error(f" 2. You have write permissions to the file")
                st.error(f" 3. No other process is using the file")
                st.error(f" Error details: {str(pe)}")
                return False
       
        except Exception as error:
            if attempt < max_retries - 1:
                logging.warning(f"Error on attempt {attempt + 1}, retrying... Error: {str(error)}")
                time.sleep(retry_delay * (attempt + 1))
                continue
            else:
                st.error(f"❌ Error appending to Excel file")
                st.error(f"📁 File path: {excel_path}")
                st.error(f"🔍 Error type: {type(error).__name__}")
                st.error(f"📝 Error message: {str(error)}")
                st.error(f"💡 Please check the file path and permissions.")
                return False
   
    return False
def get_missing_reporters(df, today):
    """Get list of employees who haven't reported today"""
    if df is None or df.empty:
        return []
    today_str = today.strftime('%Y-%m-%d')
    # Filter today's submissions (create a copy to avoid modifying original)
    if 'Date' in df.columns:
        # Convert Date column to string for comparison
        df_copy = df.copy()
        df_copy['Date'] = pd.to_datetime(df_copy['Date']).dt.strftime('%Y-%m-%d')
        today_submissions = df_copy[df_copy['Date'] == today_str]
        submitted_employees = today_submissions['Name'].unique().tolist() if 'Name' in today_submissions.columns else []
    else:
        submitted_employees = []
    # Get all employees from config
    config = load_config()
    all_employees = config.get('employee_emails', [])
    # Find missing reporters
    missing = [emp for emp in all_employees if emp not in submitted_employees]
    return missing
#Dashboard Functions
def show_metrics(df):
    """Display key metrics"""
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        total_submissions = len(df) if df is not None and not df.empty else 0
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{total_submissions}</div>
            <div class="metric-label">Total Submissions</div>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        today = datetime.now().date()
        today_count = 0
        if df is not None and not df.empty and 'Date' in df.columns:
            today_count = len(df[df['Date'] == str(today)])
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{today_count}</div>
            <div class="metric-label">Today's Reports</div>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        unique_employees = 0
        if df is not None and not df.empty and 'Name' in df.columns:
            unique_employees = df['Name'].nunique()
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{unique_employees}</div>
            <div class="metric-label">Active Employees</div>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        completed_tasks = 0
        if df is not None and not df.empty and 'Task Status' in df.columns:
            completed_tasks = len(df[df['Task Status'] == 'Completed'])
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{completed_tasks}</div>
            <div class="metric-label">Completed Tasks</div>
        </div>
        """, unsafe_allow_html=True)
def show_filters(df):
    """Display filter options"""
    if df is None or df.empty:
        return df
    st.markdown('<div class="filter-container">', unsafe_allow_html=True)
    st.subheader("🔍 Filters")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if 'Name' in df.columns:
            unique_vals = [x for x in df['Name'].unique() if pd.notna(x)]
            employees = ['All'] + sorted(unique_vals, key=lambda x: (0, x) if isinstance(x, (int, float)) else (1, str(x)))
        else:
            employees = ['All']
        selected_employee = st.selectbox("Employee", employees)
    with col2:
        if 'Project Name' in df.columns:
            unique_vals = [x for x in df['Project Name'].unique() if pd.notna(x)]
            projects = ['All'] + sorted(unique_vals, key=lambda x: (0, x) if isinstance(x, (int, float)) else (1, str(x)))
        else:
            projects = ['All']
        selected_project = st.selectbox("Project", projects)
    with col3:
        if 'Task Status' in df.columns:
            unique_vals = [x for x in df['Task Status'].unique() if pd.notna(x)]
            statuses = ['All'] + sorted(unique_vals, key=lambda x: (0, x) if isinstance(x, (int, float)) else (1, str(x)))
        else:
            statuses = ['All']
        selected_status = st.selectbox("Status", statuses)
    with col4:
        if 'Task Priority' in df.columns:
            unique_vals = [x for x in df['Task Priority'].unique() if pd.notna(x)]
            priorities = ['All'] + sorted(unique_vals, key=lambda x: (0, x) if isinstance(x, (int, float)) else (1, str(x)))
        else:
            priorities = ['All']
        selected_priority = st.selectbox("Priority", priorities)
    # Date range
    col5, col6 = st.columns(2)
    with col5:
        start_date = st.date_input("Start Date", (datetime.now() - timedelta(days=7)).date())
    with col6:
        end_date = st.date_input("End Date", datetime.now().date())
    st.markdown('</div>', unsafe_allow_html=True)
    # Apply filters
    filtered_df = df.copy()
    if 'Date' in filtered_df.columns:
        filtered_df['Date'] = pd.to_datetime(filtered_df['Date'])
        filtered_df = filtered_df[
            (filtered_df['Date'].dt.date >= start_date) &
            (filtered_df['Date'].dt.date <= end_date)
        ]
    if selected_employee != 'All' and 'Name' in filtered_df.columns:
        filtered_df = filtered_df[filtered_df['Name'] == selected_employee]
    if selected_project != 'All' and 'Project Name' in filtered_df.columns:
        filtered_df = filtered_df[filtered_df['Project Name'] == selected_project]
    if selected_status != 'All' and 'Task Status' in filtered_df.columns:
        filtered_df = filtered_df[filtered_df['Task Status'] == selected_status]
    if selected_priority != 'All' and 'Task Priority' in filtered_df.columns:
        filtered_df = filtered_df[filtered_df['Task Priority'] == selected_priority]
    return filtered_df
def show_charts(df):
    """Display analytics charts using ECharts with smooth animations"""
    if df is None or df.empty:
        st.info("No data available for charts")
        return
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📈 Task Status Distribution")
        if 'Task Status' in df.columns:
            status_counts = df['Task Status'].value_counts()
            
            # ECharts Pie Chart with animation
            pie_option = {
                "tooltip": {
                    "trigger": "item",
                    "formatter": "{b}: {c} ({d}%)",
                    "backgroundColor": "rgba(0,0,0,0.8)",
                    "borderColor": "#667eea",
                    "borderWidth": 1,
                    "textStyle": {"color": "#fff", "fontSize": 14}
                },
                "legend": {
                    "orient": "vertical",
                    "left": "left",
                    "textStyle": {"color": "#e6eef2", "fontSize": 12}
                },
                "series": [{
                    "name": "Task Status",
                    "type": "pie",
                    "radius": ["40%", "70%"],
                    "avoidLabelOverlap": False,
                    "itemStyle": {
                        "borderRadius": 10,
                        "borderColor": "#000",
                        "borderWidth": 2
                    },
                    "label": {
                        "show": True,
                        "formatter": "{b}\n{d}%",
                        "color": "#e6eef2",
                        "fontSize": 12
                    },
                    "emphasis": {
                        "label": {
                            "show": True,
                            "fontSize": 16,
                            "fontWeight": "bold"
                        },
                        "itemStyle": {
                            "shadowBlur": 10,
                            "shadowOffsetX": 0,
                            "shadowColor": "rgba(102, 126, 234, 0.5)"
                        }
                    },
                    "labelLine": {
                        "show": True,
                        "lineStyle": {"color": "#e6eef2"}
                    },
                    "data": [
                        {"value": int(count), "name": status, 
                         "itemStyle": {"color": ["#10b981", "#3b82f6", "#f59e0b", "#ef4444"][i % 4]}}
                        for i, (status, count) in enumerate(status_counts.items())
                    ],
                    "animationType": "scale",
                    "animationEasing": "elasticOut"
                }]
            }
            st_echarts(options=pie_option, height="400px", key="status_pie")
    
    with col2:
        st.subheader("⚡ Priority Distribution")
        if 'Task Priority' in df.columns:
            priority_counts = df['Task Priority'].value_counts()
            
            color_map = {
                'Low': '#10b981',
                'Medium': '#f59e0b',
                'High': '#ff6347',
                'Critical': '#dc2626'
            }
            
            # ECharts Bar Chart with animation
            bar_option = {
                "tooltip": {
                    "trigger": "axis",
                    "axisPointer": {"type": "shadow"},
                    "backgroundColor": "rgba(0,0,0,0.8)",
                    "borderColor": "#667eea",
                    "borderWidth": 1,
                    "textStyle": {"color": "#fff", "fontSize": 14}
                },
                "grid": {
                    "left": "3%",
                    "right": "4%",
                    "bottom": "3%",
                    "containLabel": True
                },
                "xAxis": [{
                    "type": "category",
                    "data": list(priority_counts.index),
                    "axisTick": {"alignWithLabel": True},
                    "axisLabel": {"color": "#e6eef2", "fontSize": 12},
                    "axisLine": {"lineStyle": {"color": "#e6eef2"}}
                }],
                "yAxis": [{
                    "type": "value",
                    "axisLabel": {"color": "#e6eef2", "fontSize": 12},
                    "axisLine": {"lineStyle": {"color": "#e6eef2"}},
                    "splitLine": {"lineStyle": {"color": "rgba(230, 238, 242, 0.1)"}}
                }],
                "series": [{
                    "name": "Tasks",
                    "type": "bar",
                    "barWidth": "60%",
                    "data": [
                        {
                            "value": int(count),
                            "itemStyle": {
                                "color": color_map.get(priority, "#667eea"),
                                "borderRadius": [10, 10, 0, 0]
                            }
                        }
                        for priority, count in priority_counts.items()
                    ],
                    "label": {
                        "show": True,
                        "position": "top",
                        "color": "#e6eef2",
                        "fontSize": 14,
                        "fontWeight": "bold"
                    },
                    "emphasis": {
                        "itemStyle": {
                            "shadowBlur": 10,
                            "shadowOffsetX": 0,
                            "shadowColor": "rgba(102, 126, 234, 0.5)"
                        }
                    }
                }]
            }
            st_echarts(options=bar_option, height="400px", key="priority_bar")
    
    # Weekly trend
    st.subheader("📊 Weekly Submission Trend")
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'])
        daily_counts = df.groupby(df['Date'].dt.date).size().reset_index(name='count')
        
        # ECharts Line Chart with smooth animations
        line_option = {
            "tooltip": {
                "trigger": "axis",
                "backgroundColor": "rgba(0,0,0,0.8)",
                "borderColor": "#667eea",
                "borderWidth": 1,
                "textStyle": {"color": "#fff", "fontSize": 14},
                "axisPointer": {
                    "type": "cross",
                    "label": {"backgroundColor": "#667eea"}
                }
            },
            "legend": {
                "data": ["Submissions"],
                "textStyle": {"color": "#e6eef2", "fontSize": 12}
            },
            "grid": {
                "left": "3%",
                "right": "4%",
                "bottom": "3%",
                "containLabel": True
            },
            "xAxis": [{
                "type": "category",
                "boundaryGap": False,
                "data": [str(d) for d in daily_counts['Date']],
                "axisLabel": {"color": "#e6eef2", "fontSize": 10, "rotate": 45},
                "axisLine": {"lineStyle": {"color": "#e6eef2"}}
            }],
            "yAxis": [{
                "type": "value",
                "axisLabel": {"color": "#e6eef2", "fontSize": 12},
                "axisLine": {"lineStyle": {"color": "#e6eef2"}},
                "splitLine": {"lineStyle": {"color": "rgba(230, 238, 242, 0.1)"}}
            }],
            "series": [{
                "name": "Submissions",
                "type": "line",
                "smooth": True,
                "symbol": "circle",
                "symbolSize": 8,
                "lineStyle": {
                    "width": 3,
                    "color": {
                        "type": "linear",
                        "x": 0, "y": 0, "x2": 1, "y2": 0,
                        "colorStops": [
                            {"offset": 0, "color": "#667eea"},
                            {"offset": 1, "color": "#764ba2"}
                        ]
                    }
                },
                "itemStyle": {"color": "#764ba2"},
                "areaStyle": {
                    "color": {
                        "type": "linear",
                        "x": 0, "y": 0, "x2": 0, "y2": 1,
                        "colorStops": [
                            {"offset": 0, "color": "rgba(102, 126, 234, 0.3)"},
                            {"offset": 1, "color": "rgba(118, 75, 162, 0.1)"}
                        ]
                    }
                },
                "emphasis": {
                    "itemStyle": {
                        "shadowBlur": 10,
                        "shadowOffsetX": 0,
                        "shadowColor": "rgba(102, 126, 234, 0.5)"
                    }
                },
                "data": [int(x) for x in daily_counts['count']]
            }]
        }
        st_echarts(options=line_option, height="400px", key="submissions_line")
def get_status_color_and_label(availability):
    """Return status label and color based on availability status"""
    if availability == "Underutilized":
        return "🟢 Underutilized", "#15b982"
    elif availability == "Partially Busy":
        return "🟡 Partially Busy", "#f59e0b"
    elif availability == "Fully Busy":
        return "🔴 Fully Busy", "#ef4444"
    else:
        return "⚪ Unknown", "#6b7280"
def format_availability_for_csv(availability):
    """Format availability value to emoji + label for CSV exports."""
    try:
        if availability is None:
            return "⚪ Unknown"
        a = str(availability).strip()
        if a == "Underutilized":
            return "🟢 Underutilized"
        if a == "Partially Busy":
            return "🟡 Partially Busy"
        if a == "Fully Busy":
            return "🔴 Fully Busy"
        return "⚪ Unknown"
    except Exception:
        return "⚪ Unknown"


def df_to_colored_excel_bytes(df: pd.DataFrame) -> bytes:
    """Return an XLSX file (bytes) with Availability column cells colour-filled.

    The function writes the DataFrame to an in-memory Excel file using pandas/openpyxl,
    then applies PatternFill colors to the Availability column based on known values.
    """
    out = io.BytesIO()
    # Use pandas to write initial sheet
    try:
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            # writer.save() is handled by context manager
    except Exception:
        # Fallback: attempt a direct openpyxl workbook creation
        out = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        # write headers
        headers = list(df.columns)
        ws.append(headers)
        for _, row in df.iterrows():
            ws.append([row.get(h, '') for h in headers])
        wb.save(out)
        out.seek(0)

    out.seek(0)
    wb = load_workbook(out)
    ws = wb.active

    # Find Availability column index
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        avail_idx = headers.index('Availability') + 1
    except ValueError:
        avail_idx = None

    # Define fills (hex without '#')
    fills = {
        'Underutilized': PatternFill(fill_type='solid', start_color='10B981', end_color='15b982'),
        'Partially Busy': PatternFill(fill_type='solid', start_color='F59E0B', end_color='f59e0b'),
        'Fully Busy': PatternFill(fill_type='solid', start_color='EF4444', end_color='ef4444'),
        'Unknown': PatternFill(fill_type='solid', start_color='6B7280', end_color='6b7280')
    }

    if avail_idx:
        for row in ws.iter_rows(min_row=2, min_col=avail_idx, max_col=avail_idx):
            cell = row[0]
            val = '' if cell.value is None else str(cell.value).strip()
            # Map common exported formats (emoji+label) back to base labels
            if 'Underutilized' in val:
                cell.fill = fills['Underutilized']
            elif 'Partially' in val:
                cell.fill = fills['Partially Busy']
            elif 'Fully' in val:
                cell.fill = fills['Fully Busy']
            else:
                cell.fill = fills['Unknown']

    # Save workbook to bytes
    out2 = io.BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2.getvalue()
def show_employee_dashboard(df):
    """Interactive dashboard for selected employee using performance metrics."""
    if df is None or df.empty or 'Name' not in df.columns:
        st.info("No employee data available for detailed view.")
        return
    df = ensure_numeric_columns(df)
    df = df.copy()
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    employees = sorted([name for name in df['Name'].dropna().unique() if str(name).strip()])
    if not employees:
        st.info("No employees found in the dataset.")
        return
    st.subheader("👨‍💻 Employee Performance Explorer")
    st.caption("Analyze and track employee performance metrics")
    # Per-employee average performance chart
    try:
        perf_summary = (
            df.groupby('Name')['Employee Performance (%)']
            .mean()
            .reset_index(name='AvgPerformance')
        )
        # latest availability per employee
        latest_avails = {}
        for name in perf_summary['Name'].tolist():
            emp_rows = df[df['Name'] == name]
            if 'Availability' in emp_rows.columns and not emp_rows[emp_rows['Availability'].notna()].empty:
                latest_avails[name] = emp_rows[emp_rows['Availability'].notna()]['Availability'].iloc[-1]
            else:
                latest_avails[name] = 'Unknown'
        perf_summary['StatusCategory'] = perf_summary['Name'].map(latest_avails)
        color_map = {
            'Underutilized': '#10b981',
            'Partially Busy': '#f59e0b',
            'Fully Busy': '#ef4444',
            'Unknown': '#6b7280'
        }
        # Use StatusCategory for colors
        perf_summary = perf_summary.sort_values('AvgPerformance', ascending=False)
        if not perf_summary.empty:
            # ECharts Bar Chart for Employee Performance
            perf_option = {
                "tooltip": {
                    "trigger": "axis",
                    "axisPointer": {"type": "shadow"},
                    "backgroundColor": "rgba(0,0,0,0.8)",
                    "borderColor": "#667eea",
                    "borderWidth": 1,
                    "textStyle": {"color": "#fff", "fontSize": 14}
                },
                "legend": {
                    "data": list(color_map.keys()),
                    "textStyle": {"color": "#e6eef2", "fontSize": 12},
                    "top": "5%"
                },
                "grid": {
                    "left": "3%",
                    "right": "4%",
                    "bottom": "15%",
                    "top": "15%",
                    "containLabel": True
                },
                "xAxis": [{
                    "type": "category",
                    "data": perf_summary['Name'].tolist(),
                    "axisTick": {"alignWithLabel": True},
                    "axisLabel": {
                        "color": "#e6eef2",
                        "fontSize": 11,
                        "rotate": 30,
                        "interval": 0
                    },
                    "axisLine": {"lineStyle": {"color": "#e6eef2"}}
                }],
                "yAxis": [{
                    "type": "value",
                    "max": 100,
                    "axisLabel": {
                        "color": "#e6eef2",
                        "fontSize": 12,
                        "formatter": "{value}%"
                    },
                    "axisLine": {"lineStyle": {"color": "#e6eef2"}},
                    "splitLine": {"lineStyle": {"color": "rgba(230, 238, 242, 0.1)"}}
                }],
                "series": [{
                    "name": status,
                    "type": "bar",
                    "barWidth": "50%",
                    "data": [
                        {
                            "value": float(row['AvgPerformance']),
                            "itemStyle": {
                                "color": color,
                                "borderRadius": [8, 8, 0, 0]
                            }
                        } if row['StatusCategory'] == status else 0
                        for _, row in perf_summary.iterrows()
                    ],
                    "label": {
                        "show": True,
                        "position": "top",
                        "color": "#e6eef2",
                        "fontSize": 11,
                        "fontWeight": "bold"
                    },
                    "emphasis": {
                        "itemStyle": {
                            "shadowBlur": 15,
                            "shadowOffsetX": 0,
                            "shadowColor": "rgba(102, 126, 234, 0.6)"
                        }
                    }
                } for status, color in color_map.items()]
            }
            st_echarts(options=perf_option, height="400px", key="employee_perf_bar")
    except Exception:
        # don't break dashboard if chart fails
        pass
    # Export All Employees (create ZIP of per-employee CSVs)
    exp_col1, exp_col2 = st.columns([5, 1])
    with exp_col2:
        if st.button("📥 Export All", use_container_width=True, key="export_all_btn"):
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                for name in employees:
                    emp_rows = df[df['Name'] == name].copy()
                    if 'Date' in emp_rows.columns:
                        emp_rows['Date'] = emp_rows['Date'].astype(str)
                    # Ensure Availability is formatted in exported CSVs
                    if 'Availability' in emp_rows.columns:
                        emp_rows_export = emp_rows.copy()
                        emp_rows_export['Availability'] = emp_rows_export['Availability'].apply(format_availability_for_csv)
                    else:
                        emp_rows_export = emp_rows
                    csv_bytes = emp_rows_export.to_csv(index=False).encode('utf-8-sig')
                    safe_name = re.sub(r"[^A-Za-z0-9_\- ]+", "", str(name)).strip() or "employee"
                    zf.writestr(f"{safe_name}_report.csv", csv_bytes)
            buf.seek(0)
            st.download_button(
                label="Download All Employees (ZIP)",
                data=buf,
                file_name=f"all_employees_reports_{datetime.now().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                key="download_all_zip"
            )
    selected_employee = st.selectbox("Select an employee to view detailed performance", ["All"] + employees, key="employee_selector")
    
    if selected_employee == "All" or not selected_employee:
        st.info("Select a specific employee to view their detailed dashboard.")
        return
    
    emp_df = df[df['Name'] == selected_employee].copy()
    if emp_df.empty:
        st.warning("No records found for the selected employee.")
        return
    total_tasks = len(emp_df)
    completed_tasks = int((emp_df.get('Task Status') == 'Completed').sum()) if 'Task Status' in emp_df.columns else 0
    pending_tasks = max(total_tasks - completed_tasks, 0)
    avg_performance = round(emp_df['Employee Performance (%)'].mean(), 2)
    latest_perf = round(emp_df.sort_values('Date')['Employee Performance (%)'].iloc[-1], 2) if not emp_df['Employee Performance (%)'].empty else 0
    last_update = emp_df['Date'].dropna().max().date().isoformat() if 'Date' in emp_df.columns and not emp_df['Date'].dropna().empty else "N/A"
   
    # Calculate additional metrics for professional display
    completion_rate = round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 1)
    productivity_score = round(avg_performance, 1)
    quality_score = round(min(avg_performance * 1.1, 100), 1)
    efficiency_score = round(min(avg_performance * 0.95, 100), 1)
   
    # Get status based on most recent Availability field from CSV
    latest_availability = emp_df[emp_df['Availability'].notna()]['Availability'].iloc[-1] if 'Availability' in emp_df.columns and not emp_df[emp_df['Availability'].notna()].empty else "Unknown"
    status_label, status_color = get_status_color_and_label(latest_availability)
   
    # Get department/project
    project_name = emp_df['Project Name'].mode()[0] if 'Project Name' in emp_df.columns and not emp_df['Project Name'].empty else 'Multiple Projects'
   
    # Professional Employee Card Display
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 25px; border-radius: 15px; margin-bottom: 20px;
                box-shadow: 0 4px 15px rgba(0,0,0,0.2);">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
            <div>
                <h2 style="color: white; margin: 0; font-size: 28px;">{selected_employee}</h2>
                <p style="color: rgba(255,255,255,0.9); margin: 5px 0 0 0; font-size: 16px;">
                    {project_name}
                </p>
            </div>
            <div style="text-align: right;">
                <span style="background: {status_color}; color: white; padding: 8px 16px;
                             border-radius: 20px; font-weight: bold; font-size: 14px;">
                    {status_label}
                </span>
            </div>
        </div>
        <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; margin-top: 15px;">
            <div style="text-align: center;">
                <p style="color: rgba(255,255,255,0.8); margin: 0; font-size: 13px;">Productivity</p>
                <p style="color: white; margin: 5px 0 0 0; font-size: 24px; font-weight: bold;">{productivity_score}%</p>
            </div>
            <div style="text-align: center;">
                <p style="color: rgba(255,255,255,0.8); margin: 0; font-size: 13px;">Quality</p>
                <p style="color: white; margin: 5px 0 0 0; font-size: 24px; font-weight: bold;">{quality_score}%</p>
            </div>
            <div style="text-align: center;">
                <p style="color: rgba(255,255,255,0.8); margin: 0; font-size: 13px;">Efficiency</p>
                <p style="color: white; margin: 5px 0 0 0; font-size: 24px; font-weight: bold;">{efficiency_score}%</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
   
    # Export individual employee data
    col_export_individual = st.columns([5, 1])
    with col_export_individual[1]:
        # Prepare export data
        export_df = emp_df.copy()
        if 'Date' in export_df.columns:
            export_df['Date'] = export_df['Date'].astype(str)
        export_df = export_df.sort_values('Date', ascending=False) if 'Date' in export_df.columns else export_df
        # Format Availability for CSV export
        if 'Availability' in export_df.columns:
            export_df = export_df.copy()
            export_df['Availability'] = export_df['Availability'].apply(format_availability_for_csv)
        csv_bytes = export_df.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label=f"📥 Export",
            data=csv_bytes,
            file_name=f"{selected_employee}_performance_report_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
            key="export_individual_emp"
        )
        try:
            excel_bytes = df_to_colored_excel_bytes(export_df)
            st.download_button(
                label=f"📥 Export Excel",
                data=excel_bytes,
                file_name=f"{selected_employee}_performance_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="export_individual_emp_xlsx"
            )
        except Exception:
            # Non-fatal: if excel generation fails, keep CSV available
            pass
    metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
    with metric_col1:
        st.metric("Total Tasks", total_tasks)
    with metric_col2:
        st.metric("Completed Tasks", completed_tasks)
    with metric_col3:
        st.metric("Completion Rate", f"{completion_rate}%")
    with metric_col4:
        st.metric("Last Update", last_update)
    chart_col1, chart_col2 = st.columns([1, 1])
    with chart_col1:
        st.caption("Current Performance Gauge")
        # ECharts Gauge with smooth animation
        gauge_option = {
            "tooltip": {
                "formatter": "{b}: {c}%",
                "backgroundColor": "rgba(0,0,0,0.8)",
                "borderColor": "#667eea",
                "textStyle": {"color": "#fff", "fontSize": 14}
            },
            "series": [{
                "type": "gauge",
                "startAngle": 180,
                "endAngle": 0,
                "center": ["50%", "75%"],
                "radius": "90%",
                "min": 0,
                "max": 100,
                "splitNumber": 10,
                "axisLine": {
                    "lineStyle": {
                        "width": 6,
                        "color": [
                            [0.5, "#ff7675"],
                            [0.8, "#ffeaa7"],
                            [1, "#55efc4"]
                        ]
                    }
                },
                "pointer": {
                    "icon": "path://M12.8,0.7l12,40.1H0.7L12.8,0.7z",
                    "length": "12%",
                    "width": 20,
                    "offsetCenter": [0, "-60%"],
                    "itemStyle": {"color": "#764ba2"}
                },
                "axisTick": {
                    "length": 12,
                    "lineStyle": {"color": "auto", "width": 2}
                },
                "splitLine": {
                    "length": 20,
                    "lineStyle": {"color": "auto", "width": 5}
                },
                "axisLabel": {
                    "color": "#e6eef2",
                    "fontSize": 12,
                    "distance": -60,
                    "formatter": "{value}%"
                },
                "title": {
                    "offsetCenter": [0, "-10%"],
                    "fontSize": 14,
                    "color": "#e6eef2"
                },
                "detail": {
                    "fontSize": 28,
                    "offsetCenter": [0, "-35%"],
                    "valueAnimation": True,
                    "formatter": "{value}%",
                    "color": "#764ba2"
                },
                "data": [{
                    "value": float(latest_perf),
                    "name": "Latest Performance"
                }]
            }]
        }
        st_echarts(options=gauge_option, height="300px", key=f"gauge_{selected_employee}")
    
    with chart_col2:
        st.caption("Performance Snapshot")
        trend_df = emp_df[['Date', 'Employee Performance (%)']].dropna()
        if not trend_df.empty and trend_df['Date'].notna().any():
            trend_df = trend_df.sort_values('Date')
            
            # Format dates safely
            date_labels = []
            for d in trend_df['Date']:
                if isinstance(d, str):
                    date_labels.append(d)
                elif hasattr(d, 'date'):
                    date_labels.append(str(d.date()))
                else:
                    date_labels.append(str(d))
            
            # ECharts Line Chart
            snapshot_option = {
                "tooltip": {
                    "trigger": "axis",
                    "backgroundColor": "rgba(0,0,0,0.8)",
                    "borderColor": "#667eea",
                    "textStyle": {"color": "#fff", "fontSize": 14}
                },
                "grid": {
                    "left": "10%",
                    "right": "5%",
                    "bottom": "15%",
                    "top": "10%"
                },
                "xAxis": {
                    "type": "category",
                    "data": date_labels,
                    "axisLabel": {
                        "color": "#e6eef2",
                        "fontSize": 10,
                        "rotate": 30
                    },
                    "axisLine": {"lineStyle": {"color": "#e6eef2"}}
                },
                "yAxis": {
                    "type": "value",
                    "min": 0,
                    "max": 100,
                    "axisLabel": {
                        "color": "#e6eef2",
                        "fontSize": 11,
                        "formatter": "{value}%"
                    },
                    "axisLine": {"lineStyle": {"color": "#e6eef2"}},
                    "splitLine": {"lineStyle": {"color": "rgba(230, 238, 242, 0.1)"}}
                },
                "series": [{
                    "type": "line",
                    "smooth": True,
                    "symbol": "circle",
                    "symbolSize": 8,
                    "data": [float(x) for x in trend_df['Employee Performance (%)']],
                    "lineStyle": {
                        "width": 3,
                        "color": "#764ba2"
                    },
                    "itemStyle": {"color": "#764ba2"},
                    "areaStyle": {
                        "color": {
                            "type": "linear",
                            "x": 0, "y": 0, "x2": 0, "y2": 1,
                            "colorStops": [
                                {"offset": 0, "color": "rgba(118, 75, 162, 0.3)"},
                                {"offset": 1, "color": "rgba(118, 75, 162, 0.05)"}
                            ]
                        }
                    },
                    "emphasis": {
                        "itemStyle": {
                            "shadowBlur": 10,
                            "shadowColor": "rgba(118, 75, 162, 0.5)"
                        }
                    }
                }]
            }
            st_echarts(options=snapshot_option, height="300px", key=f"snapshot_{selected_employee}")
        else:
            st.info("No performance history available for this employee.")
    st.caption("Task Breakdown")
    breakdown_col1, breakdown_col2 = st.columns(2)
    with breakdown_col1:
        if 'Task Status' in emp_df.columns:
            status_counts = emp_df['Task Status'].value_counts()
            if not status_counts.empty:
                # ECharts Pie for Task Status
                status_pie_option = {
                    "tooltip": {
                        "trigger": "item",
                        "formatter": "{b}: {c} ({d}%)",
                        "backgroundColor": "rgba(0,0,0,0.8)",
                        "borderColor": "#667eea",
                        "textStyle": {"color": "#fff", "fontSize": 14}
                    },
                    "legend": {
                        "orient": "horizontal",
                        "bottom": "0%",
                        "textStyle": {"color": "#e6eef2", "fontSize": 11}
                    },
                    "series": [{
                        "name": "Task Status",
                        "type": "pie",
                        "radius": ["30%", "60%"],
                        "center": ["50%", "45%"],
                        "avoidLabelOverlap": True,
                        "itemStyle": {
                            "borderRadius": 8,
                            "borderColor": "#000",
                            "borderWidth": 2
                        },
                        "label": {
                            "show": True,
                            "formatter": "{d}%",
                            "color": "#e6eef2",
                            "fontSize": 12
                        },
                        "emphasis": {
                            "label": {
                                "show": True,
                                "fontSize": 14,
                                "fontWeight": "bold"
                            },
                            "itemStyle": {
                                "shadowBlur": 10,
                                "shadowOffsetX": 0,
                                "shadowColor": "rgba(102, 126, 234, 0.5)"
                            }
                        },
                        "data": [
                            {
                                "value": int(count),
                                "name": status,
                                "itemStyle": {
                                    "color": ["#3b82f6", "#10b981", "#f59e0b", "#ef4444"][i % 4]
                                }
                            }
                            for i, (status, count) in enumerate(status_counts.items())
                        ],
                        "animationType": "scale",
                        "animationEasing": "elasticOut"
                    }]
                }
                st_echarts(options=status_pie_option, height="300px", key=f"status_pie_{selected_employee}")
            else:
                st.info("No task status data available for this employee.")
        else:
            st.info("Task status column not available.")
    
    with breakdown_col2:
        if 'Task Priority' in emp_df.columns:
            priority_counts = emp_df['Task Priority'].value_counts()
            if not priority_counts.empty:
                priority_color_map = {
                    'Low': '#10b981',
                    'Medium': '#f59e0b',
                    'High': '#ff6347',
                    'Critical': '#dc2626'
                }
                # ECharts Bar for Priority
                priority_bar_option = {
                    "tooltip": {
                        "trigger": "axis",
                        "axisPointer": {"type": "shadow"},
                        "backgroundColor": "rgba(0,0,0,0.8)",
                        "borderColor": "#667eea",
                        "textStyle": {"color": "#fff", "fontSize": 14}
                    },
                    "grid": {
                        "left": "15%",
                        "right": "5%",
                        "bottom": "10%",
                        "top": "10%"
                    },
                    "xAxis": {
                        "type": "category",
                        "data": list(priority_counts.index),
                        "axisLabel": {
                            "color": "#e6eef2",
                            "fontSize": 11
                        },
                        "axisLine": {"lineStyle": {"color": "#e6eef2"}}
                    },
                    "yAxis": {
                        "type": "value",
                        "axisLabel": {
                            "color": "#e6eef2",
                            "fontSize": 11
                        },
                        "axisLine": {"lineStyle": {"color": "#e6eef2"}},
                        "splitLine": {"lineStyle": {"color": "rgba(230, 238, 242, 0.1)"}}
                    },
                    "series": [{
                        "name": "Tasks",
                        "type": "bar",
                        "data": [
                            {
                                "value": int(count),
                                "itemStyle": {
                                    "color": priority_color_map.get(priority, "#667eea"),
                                    "borderRadius": [8, 8, 0, 0]
                                }
                            }
                            for priority, count in priority_counts.items()
                        ],
                        "label": {
                            "show": True,
                            "position": "top",
                            "color": "#e6eef2",
                            "fontSize": 13,
                            "fontWeight": "bold"
                        },
                        "emphasis": {
                            "itemStyle": {
                                "shadowBlur": 10,
                                "shadowColor": "rgba(102, 126, 234, 0.5)"
                            }
                        }
                    }]
                }
                st_echarts(options=priority_bar_option, height="300px", key=f"priority_bar_{selected_employee}")
            else:
                st.info("No task priority data available for this employee.")
        else:
            st.info("Task priority column not available.")
    st.subheader("📈 Performance Trend")
    st.caption("Track performance metrics over time")
   
    trend_df = emp_df[['Date', 'Employee Performance (%)']].dropna()
    if not trend_df.empty and trend_df['Date'].notna().any():
        # Add productivity, quality, efficiency calculations for trend
        trend_df = trend_df.sort_values('Date')
        trend_df['Productivity'] = trend_df['Employee Performance (%)']
        trend_df['Quality'] = (trend_df['Employee Performance (%)'] * 1.1).clip(upper=100)
        trend_df['Efficiency'] = (trend_df['Employee Performance (%)'] * 0.95).clip(upper=100)
       
        # Format dates safely
        date_labels = []
        for d in trend_df['Date']:
            if isinstance(d, str):
                date_labels.append(d)
            elif hasattr(d, 'date'):
                date_labels.append(str(d.date()))
            else:
                date_labels.append(str(d))
       
        # ECharts Multi-line Trend Chart
        trend_option = {
            "title": {
                "text": f"{selected_employee}'s Performance Trend",
                "textStyle": {"color": "#e6eef2", "fontSize": 16},
                "left": "center"
            },
            "tooltip": {
                "trigger": "axis",
                "backgroundColor": "rgba(0,0,0,0.8)",
                "borderColor": "#667eea",
                "borderWidth": 1,
                "textStyle": {"color": "#fff", "fontSize": 14},
                "axisPointer": {
                    "type": "cross",
                    "label": {"backgroundColor": "#667eea"}
                }
            },
            "legend": {
                "data": ["Productivity", "Quality", "Efficiency"],
                "textStyle": {"color": "#e6eef2", "fontSize": 12},
                "top": "8%"
            },
            "grid": {
                "left": "5%",
                "right": "5%",
                "bottom": "15%",
                "top": "20%",
                "containLabel": True
            },
            "xAxis": {
                "type": "category",
                "boundaryGap": False,
                "data": date_labels,
                "axisLabel": {
                    "color": "#e6eef2",
                    "fontSize": 10,
                    "rotate": 30
                },
                "axisLine": {"lineStyle": {"color": "#e6eef2"}}
            },
            "yAxis": {
                "type": "value",
                "min": 0,
                "max": 100,
                "axisLabel": {
                    "color": "#e6eef2",
                    "fontSize": 11,
                    "formatter": "{value}%"
                },
                "axisLine": {"lineStyle": {"color": "#e6eef2"}},
                "splitLine": {"lineStyle": {"color": "rgba(230, 238, 242, 0.1)"}}
            },
            "series": [
                {
                    "name": "Productivity",
                    "type": "line",
                    "smooth": True,
                    "symbol": "circle",
                    "symbolSize": 8,
                    "data": [float(x) for x in trend_df['Productivity']],
                    "lineStyle": {
                        "width": 3,
                        "color": "#3b82f6"
                    },
                    "itemStyle": {"color": "#3b82f6"},
                    "areaStyle": {
                        "color": {
                            "type": "linear",
                            "x": 0, "y": 0, "x2": 0, "y2": 1,
                            "colorStops": [
                                {"offset": 0, "color": "rgba(59, 130, 246, 0.3)"},
                                {"offset": 1, "color": "rgba(59, 130, 246, 0.05)"}
                            ]
                        }
                    },
                    "emphasis": {
                        "itemStyle": {
                            "shadowBlur": 10,
                            "shadowColor": "rgba(59, 130, 246, 0.5)"
                        }
                    }
                },
                {
                    "name": "Quality",
                    "type": "line",
                    "smooth": True,
                    "symbol": "circle",
                    "symbolSize": 8,
                    "data": [float(x) for x in trend_df['Quality']],
                    "lineStyle": {
                        "width": 3,
                        "color": "#10b981"
                    },
                    "itemStyle": {"color": "#10b981"},
                    "areaStyle": {
                        "color": {
                            "type": "linear",
                            "x": 0, "y": 0, "x2": 0, "y2": 1,
                            "colorStops": [
                                {"offset": 0, "color": "rgba(16, 185, 129, 0.3)"},
                                {"offset": 1, "color": "rgba(16, 185, 129, 0.05)"}
                            ]
                        }
                    },
                    "emphasis": {
                        "itemStyle": {
                            "shadowBlur": 10,
                            "shadowColor": "rgba(16, 185, 129, 0.5)"
                        }
                    }
                },
                {
                    "name": "Efficiency",
                    "type": "line",
                    "smooth": True,
                    "symbol": "circle",
                    "symbolSize": 8,
                    "data": [float(x) for x in trend_df['Efficiency']],
                    "lineStyle": {
                        "width": 3,
                        "color": "#f59e0b"
                    },
                    "itemStyle": {"color": "#f59e0b"},
                    "areaStyle": {
                        "color": {
                            "type": "linear",
                            "x": 0, "y": 0, "x2": 0, "y2": 1,
                            "colorStops": [
                                {"offset": 0, "color": "rgba(245, 158, 11, 0.3)"},
                                {"offset": 1, "color": "rgba(245, 158, 11, 0.05)"}
                            ]
                        }
                    },
                    "emphasis": {
                        "itemStyle": {
                            "shadowBlur": 10,
                            "shadowColor": "rgba(245, 158, 11, 0.5)"
                        }
                    }
                }
            ]
        }
        st_echarts(options=trend_option, height="450px", key=f"trend_{selected_employee}")
       
        # Performance statistics table
        st.markdown('**Performance Statistics**')
        stats_col1, stats_col2, stats_col3 = st.columns(3)
        with stats_col1:
            st.metric("Avg Productivity", f"{trend_df['Productivity'].mean():.1f}%")
        with stats_col2:
            st.metric("Avg Quality", f"{trend_df['Quality'].mean():.1f}%")
        with stats_col3:
            st.metric("Avg Efficiency", f"{trend_df['Efficiency'].mean():.1f}%")
       
        # Recent performance data with download option
        st.markdown('**Recent Performance Values**')
        recent_performance_df = trend_df[['Date', 'Productivity', 'Quality', 'Efficiency']].copy()
        recent_performance_df = recent_performance_df.sort_values('Date', ascending=False).head(20)
        recent_performance_df['Date'] = recent_performance_df['Date'].dt.strftime('%Y-%m-%d')
        st.dataframe(recent_performance_df, use_container_width=True)
       
        # Download trend data
        trend_csv_bytes = trend_df.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label="📥 Download Performance Trend Data",
            data=trend_csv_bytes,
            file_name=f"{selected_employee}_performance_trend_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            key="download_trend_data"
        )
    else:
        st.info("No performance history available for this employee.")
    st.subheader("📋 Recent Tasks")
    display_columns = [col for col in DATA_COLUMNS if col in emp_df.columns]
    if display_columns:
        display_df = emp_df.sort_values('Date', ascending=False)[display_columns]
       
        def highlight_support(val):
            if pd.isna(val) or str(val).strip() == "":
                return ""
            return "background-color: #ffcccb; color: #000000; font-weight: bold; border-left: 3px solid #ff4444"
       
        styled_df = display_df.style.map(highlight_support, subset=['Support Request'])
        st.dataframe(styled_df, use_container_width=True, height=320)
    else:
        st.info("No detailed task records to display.")
def show_data_table(df):
    """Display data table"""
    st.subheader("📋 Recent Submissions")
    if df is None or df.empty:
        st.info("No submissions found")
        return
    # Display options
    col1, col2 = st.columns([3, 1])
    with col1:
        search = st.text_input("🔎 Search", placeholder="Search in any column...")
    with col2:
        rows_to_show = st.number_input("Rows", min_value=10, max_value=1000, value=50, step=10)
    # Apply search
    display_df = df.copy()
    if search:
        mask = display_df.astype(str).apply(
            lambda x: x.str.contains(search, case=False, na=False)
        ).any(axis=1)
        display_df = display_df[mask]
    # NEW: Helper function to highlight non-empty Support Request cells
    def highlight_support(val):
        if pd.isna(val) or str(val).strip() == "":
            return ""
        return "background-color: #ffcccb; color: #000000; font-weight: bold; border-left: 3px solid #ff4444" # Light red for flagged support requests
    # NEW: Apply styling to the Support Request column only
    styled_df = display_df.style.map(highlight_support, subset=['Support Request'])
    # UPDATED: Use the styled dataframe
    st.dataframe(
        styled_df,
        use_container_width=True,
        height=400
    )
    # Download button
    if not display_df.empty:
        df_export = display_df.copy()
        # Ensure Availability column exports with emoji labels
        if 'Availability' in df_export.columns:
            df_export['Availability'] = df_export['Availability'].apply(format_availability_for_csv)
        csv_bytes = df_export.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label="📥 Download Data as CSV",
            data=csv_bytes,
            file_name=f"employee_progress_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv"
        )
        try:
            excel_bytes_all = df_to_colored_excel_bytes(df_export)
            st.download_button(
                label="📥 Download Data as Excel",
                data=excel_bytes_all,
                file_name=f"employee_progress_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception:
            pass
#Settings Page
def show_settings():
    """Display settings page"""
    st.title("⚙️ Settings")
    config = load_config()
    with st.form("settings_form"):
        st.markdown("---")
        st.subheader("Reminder Settings")
        col1, col2 = st.columns(2)
        with col1:
            reminder_time = st.time_input(
                "Reminder Time",
                value=datetime.strptime(config.get('reminder_time', '18:00'), '%H:%M').time()
            )
        with col2:
            st.write("Reminder Days (uncheck Sunday)")
            reminder_days = []
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            default_days = config.get('reminder_days', [0, 1, 2, 3, 4, 5])
            for i, day in enumerate(days):
                if st.checkbox(day, value=i in default_days, key=f"day_{i}"):
                    reminder_days.append(i)
        st.markdown("---")
        st.subheader("Email Configuration")
        admin_email = st.text_input(
            "Admin Email",
            value=config.get('admin_email', '')
        )
        employee_emails = st.text_area(
            "Employee Emails (one per line)",
            value='\n'.join(config.get('employee_emails', [])),
            height=150
        )
        st.caption("If you enabled Telegram reminders, add chat IDs in the same order as emails.")
        employee_telegram_chat_ids = st.text_area(
            "Employee Telegram Chat IDs (one per line, aligned with emails)",
            value='\n'.join([str(x) for x in config.get('employee_telegram_chat_ids', [])]),
            height=150
        )
        submitted = st.form_submit_button("💾 Save Settings")
        if submitted:
            # Update config
            config['reminder_time'] = reminder_time.strftime('%H:%M')
            config['reminder_days'] = reminder_days
            config['admin_email'] = admin_email
            config['employee_emails'] = [
                email.strip()
                for email in employee_emails.split('\n')
                if email.strip()
            ]
            # Parse Telegram chat IDs line by line, keep as int if numeric, else string
            parsed_chat_ids = []
            for line in employee_telegram_chat_ids.split('\n'):
                raw = line.strip()
                if not raw:
                    continue
                # Try int conversion (supports negative IDs)
                try:
                    parsed_chat_ids.append(int(raw))
                except ValueError:
                    parsed_chat_ids.append(raw)
            config['employee_telegram_chat_ids'] = parsed_chat_ids
            save_config(config)
            st.success("✅ Settings saved successfully!")
            time.sleep(1)
            st.rerun()

# Submit Report Page
def show_submit_report():
    """Display form for submitting work progress reports with multiple tasks"""
    config = load_config()
   
    # Logo Section - Centered at top
    # Try multiple path approaches for compatibility
    # logo_found = False
    # possible_paths = [
    # "logo/PTF1.png", # Simple relative path (usually works on Streamlit Cloud)
    # Path("logo/PTF1.png"), # Path object version
    # Path(__file__).parent / "logo" / "PTF1.png", # Relative to script
    # ]
    # --- Logo Section (Fixed for Streamlit Cloud) ---
    # Use a single HTML block to ensure the image is centered and constrained
    logo_url = "https://raw.githubusercontent.com/SoumyaR01/Employee-Task-Tracker/main/logo/ptf.png"
    # Render using an HTML <img> so CSS sizing/centering is reliable inside Streamlit's layout
    try:
        st.markdown(
            f'<div class="logo-container" style="text-align:center;">'
            f' <img src="{logo_url}" alt="PTF Logo" '
            f' style="max-width:360px; width:100%; height:auto; display:block; margin:0 auto; border-radius:8px;"/>'
            f'</div>',
            unsafe_allow_html=True,
        )
    except Exception:
        # Fallback placeholder if remote image cannot be loaded
        st.markdown(
            """
        <div class="logo-container" style="text-align: center;">
            <div style="width: 100%; max-width:360px; height: auto; padding: 12px 0; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        display: flex; align-items: center; justify-content: center;
                        border-radius: 10px; border: 2px solid #667eea; margin:0 auto;">
                <p style="color: white; font-size: 18px; font-weight: bold; margin: 0;">PTF</p>
            </div>
        </div>
        """,
            unsafe_allow_html=True,
        )
   
    # Title Section - Centered below logo
    st.markdown("<h1 style='text-align: center; margin-top: 10px; color: #2c3e50;'>Staff Productivity Insight</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #7f8c8d; font-size: 1.1rem;'>📤Submit your task report for today</p>", unsafe_allow_html=True)
   
    st.markdown("---")
    # DB-only mode: we no longer use a local Excel file path
    # Initialize session state for task count if not exists
    if 'num_tasks' not in st.session_state:
        st.session_state.num_tasks = 1
   
    st.subheader("👨‍💻 Employee Information")
   
    col1, col2 = st.columns(2)
   
    with col1:
        date = st.date_input("Date*", value=datetime.now().date())
        work_mode = st.selectbox(
            "Work Mode*",
            ["", "PTF", "Remote"],
            help="Select your work mode"
        )
   
    with col2:
        emp_id = st.text_input(
            "Employee ID*",
            placeholder="Enter your employee ID",
            help="Required field"
        )
        name = st.text_input(
            "Name*",
            placeholder="Enter your full name",
            help="Required field"
        )
   
    st.markdown("---")
    st.subheader("📋 Today's Tasks")
    st.info("💡 Record all tasks you worked on today. Multiple entries are allowed before submission.")
   
    # Handle add/remove task buttons
    col_add_remove = st.columns([1, 1, 4])
    with col_add_remove[0]:
        if st.button("➕ Add Task", use_container_width=True):
            st.session_state.num_tasks += 1
            st.rerun()
    with col_add_remove[1]:
        if st.button("➖ Remove Task", use_container_width=True) and st.session_state.num_tasks > 1:
            st.session_state.num_tasks -= 1
            st.rerun()
   
    st.caption(f"📋 You have {st.session_state.num_tasks} task(s) in this report")
   
    # Display task inputs
    for i in range(st.session_state.num_tasks):
        with st.expander(f"Task {i+1}", expanded=(i == 0)):
            col3, col4 = st.columns(2)
           
            with col3:
                project_name = st.text_input(
                    "Project Name*",
                    placeholder="Enter project name",
                    help="Required field",
                    key=f"project_{i}"
                )
                task_title = st.text_input(
                    "Task Title*",
                    placeholder="Describe the task...",
                    help="Brief description of the task",
                    key=f"title_{i}"
                )
                task_assigned_by = st.text_input(
                    "Task Assigned By*",
                    placeholder="Who assigned this task?",
                    help="Person who assigned the task",
                    key=f"assigned_{i}"
                )
           
            with col4:
                task_priority = st.selectbox(
                    "Task Priority*",
                    ["", "Low", "Medium", "High", "Critical"],
                    help="Select task priority level",
                    key=f"priority_{i}"
                )
                task_status = st.selectbox(
                    "Task Status*",
                    ["", "In Progress", "Completed"],
                    help="Select current task status",
                    key=f"status_{i}"
                )
                effort = st.number_input(
                    "Effort (in hours)*",
                    min_value=0.0,
                    value=1.0, # Default to 1 hour
                    step=0.5,
                    help="Hours spent on this task (must be >0)",
                    key=f"effort_{i}"
                )
                comments = st.text_area(
                    "Support Request",
                    placeholder="Provide any supporting information...",
                    height=80,
                    help="Optional comments",
                    key=f"comments_{i}"
                )
                availability = st.selectbox(
                    "Availability",
                    ["", "Fully Busy", "Partially Busy", "Underutilized"],
                    help="Select availability status for this task",
                    key=f"availability_{i}"
                )
   
    st.markdown("---")
    st.subheader("📅 Plan for Tomorrow")
   
    plan_for_next_day = st.text_area(
        "Plan for Next Day*",
        placeholder="What are your plans for tomorrow?",
        height=100,
        help="Required field"
    )
   
    submitted = st.button("✅ Submit Daily Report", use_container_width=True)
   
    if submitted:
        # Validate employee information
        employee_fields = {
            "Date": date,
            "Work Mode": work_mode,
            "Employee ID": emp_id,
            "Name": name
        }
       
        missing_employee_fields = [field for field, value in employee_fields.items() if not value]
       
        if missing_employee_fields:
            st.error(f"❌ Please fill in all employee information: {', '.join(missing_employee_fields)}")
        elif st.session_state.num_tasks == 0:
            st.error("❌ Please add at least one task to your report.")
        elif not plan_for_next_day:
            st.error("❌ Please fill in your plan for next day.")
        else:
            # Collect all task data from session_state (widget values are stored there with keys)
            task_data_list = []
            invalid_tasks = []
           
            for i in range(st.session_state.num_tasks):
                # Get values from session_state (widgets with keys store values there with keys)
                project_name = st.session_state.get(f"project_{i}", "")
                task_title = st.session_state.get(f"title_{i}", "")
                task_assigned_by = st.session_state.get(f"assigned_{i}", "")
                task_priority = st.session_state.get(f"priority_{i}", "")
                task_status = st.session_state.get(f"status_{i}", "")
                effort = st.session_state.get(f"effort_{i}", 0.0)
                comments = st.session_state.get(f"comments_{i}", "")
               
                # Validate task
                if not all([project_name, task_title, task_assigned_by, task_priority, task_status]) or effort <= 0:
                    invalid_tasks.append(i + 1)
                else:
                    availability = st.session_state.get(f"availability_{i}", "")
                    task_data_list.append({
                        'Date': date.strftime("%Y-%m-%d"),
                        'Work Mode': work_mode,
                        'Emp Id': emp_id,
                        'Name': name,
                        'Project Name': project_name,
                        'Task Title': task_title,
                        'Task Assigned By': task_assigned_by,
                        'Task Priority': task_priority,
                        'Task Status': task_status,
                        'Plan for next day': plan_for_next_day,
                                'Support Request': comments if comments else '',
                            'Availability': availability if availability else '',
                            'Effort (in hours)': effort,
                        # 'Employee Performance (%)' calculated below
                    })
           
            # Calculate overall performance for the day using the new priority/effort formula
            if task_data_list:
                performance = calculate_performance(task_data_list)
                for row in task_data_list:
                    row['Employee Performance (%)'] = performance
            if invalid_tasks:
                st.error(f"❌ Please fill in all required fields for task(s): {', '.join(map(str, invalid_tasks))}")
            elif not task_data_list:
                st.error("❌ No valid tasks to submit. Please add at least one complete task.")
            else:
                # Append all tasks to Excel file
                    with st.spinner(f"Saving your daily report with {len(task_data_list)} task(s)..."):
                        success = append_to_excel(task_data_list)
               
            if success:
                    st.success(f"✅ Your daily work progress report has been submitted successfully! ({len(task_data_list)} task(s) recorded)")
                    st.balloons()
                    # Reset task count for next submission
                    st.session_state.num_tasks = 1
                    # Clear form values by clearing session state keys
                    for i in range(10): # Clear up to 10 task slots
                        for key_suffix in ['project', 'title', 'assigned', 'priority', 'status', 'comments', 'availability', 'effort']:
                            key = f"{key_suffix}_{i}"
                            if key in st.session_state:
                                del st.session_state[key]
                    time.sleep(2)
                    st.rerun()
            else:
                    st.error("❌ Failed to save report. Please try again or contact administrator.")
# Main App

# --- Authentication and Routing ---
def show_login_signup():
    """Show beautiful login/signup page with modern design"""
    
    # Custom CSS for modern authentication UI
    st.markdown("""
    <style>
        /* Hide Streamlit branding */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        .stDeployButton {display: none;}
        
        /* Main background */
        .stApp {
            background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
        }
        
        /* Center container */
        .main .block-container {
            max-width: 550px;
            padding-top: 3rem;
            padding-bottom: 3rem;
        }
        
        /* Logo section */
        .auth-header {
            text-align: center;
            margin-bottom: 2.5rem;
            animation: fadeInDown 0.8s ease;
        }
        
        @keyframes fadeInDown {
            from {
                opacity: 0;
                transform: translateY(-20px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        
        .auth-title {
            font-size: 2.8rem;
            font-weight: 700;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 0.5rem;
            letter-spacing: -0.5px;
        }
        
        .auth-subtitle {
            color: #94a3b8;
            font-size: 1.1rem;
            font-weight: 400;
            margin-top: 0.5rem;
        }
        
        /* Tabs styling */
        .stTabs {
            margin-bottom: 2rem;
        }
        
        .stTabs [data-baseweb="tab-list"] {
            gap: 1rem;
            background: rgba(15, 23, 42, 0.4);
            border-radius: 12px;
            padding: 0.5rem;
            justify-content: center;
        }
        
        .stTabs [data-baseweb="tab"] {
            height: 50px;
            color: #94a3b8;
            background: transparent;
            border-radius: 8px;
            border: none;
            font-size: 1.05rem;
            font-weight: 500;
            padding: 0 2rem;
            transition: all 0.3s ease;
        }
        
        .stTabs [data-baseweb="tab"]:hover {
            background: rgba(102, 126, 234, 0.1);
            color: #667eea;
        }
        
        .stTabs [aria-selected="true"] {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white !important;
        }
        
        /* Form container */
        .form-container {
            background: rgba(15, 23, 42, 0.6);
            border-radius: 20px;
            padding: 2.5rem;
            backdrop-filter: blur(20px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.5);
            animation: fadeInUp 0.8s ease;
        }
        
        @keyframes fadeInUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        
        /* Form title */
        .form-title {
            color: #e2e8f0;
            font-size: 1.5rem;
            font-weight: 600;
            margin-bottom: 1.5rem;
            text-align: center;
        }
        
        /* Input fields */
        .stTextInput label, .stRadio label {
            color: #cbd5e1 !important;
            font-weight: 500 !important;
            font-size: 0.9rem !important;
            margin-bottom: 0.5rem !important;
        }
        
        .stTextInput > div > div > input {
            background: rgba(30, 41, 59, 0.6) !important;
            border: 1.5px solid rgba(148, 163, 184, 0.3) !important;
            border-radius: 10px !important;
            color: #e2e8f0 !important;
            padding: 0.85rem 1.2rem !important;
            font-size: 1rem !important;
            transition: all 0.3s ease !important;
        }
        
        .stTextInput > div > div > input::placeholder {
            color: #64748b !important;
        }
        
        .stTextInput > div > div > input:focus {
            border-color: #667eea !important;
            box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.15) !important;
            background: rgba(30, 41, 59, 0.8) !important;
        }
        
        /* Radio buttons */
        .stRadio > div {
            background: rgba(30, 41, 59, 0.4);
            padding: 0.8rem;
            border-radius: 10px;
            border: 1px solid rgba(148, 163, 184, 0.2);
        }
        
        .stRadio > div > label > div[data-testid="stMarkdownContainer"] p {
            color: #e2e8f0 !important;
            font-size: 0.95rem !important;
        }
        
        /* Buttons */
        .stButton button {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
            color: white !important;
            border: none !important;
            border-radius: 10px !important;
            padding: 0.85rem 2rem !important;
            font-size: 1.05rem !important;
            font-weight: 600 !important;
            width: 100% !important;
            margin-top: 1.5rem !important;
            transition: all 0.3s ease !important;
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4) !important;
        }
        
        .stButton button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 8px 25px rgba(102, 126, 234, 0.5) !important;
        }
        
        .stButton button:active {
            transform: translateY(0) !important;
        }
        
        /* Alert messages */
        .stAlert {
            border-radius: 10px !important;
            backdrop-filter: blur(10px);
        }
        
        /* Success/Error messages */
        [data-testid="stSuccess"], [data-testid="stError"], [data-testid="stInfo"] {
            border-radius: 10px !important;
        }
        
        /* Divider */
        .divider {
            height: 1px;
            background: linear-gradient(90deg, transparent, rgba(148, 163, 184, 0.3), transparent);
            margin: 2rem 0;
        }
        
        /* Footer text */
        .footer-text {
            text-align: center;
            color: #94a3b8;
            font-size: 0.9rem;
            margin-top: 1.5rem;
        }
        
        .footer-text a {
            color: #667eea;
            text-decoration: none;
            font-weight: 500;
        }
        
        .footer-text a:hover {
            text-decoration: underline;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Logo and Header
    st.markdown("""
    <div class="auth-header">
        <div class="auth-title">🔐 Employee Task Tracker</div>
        <div class="auth-subtitle">Login or signup to continue</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Tabs for User and Admin
    tab1, tab2 = st.tabs(["� User Access", "�️ Admin Access"])
    
    # TAB 1: USER ACCESS (Toggle between Login and Signup)
    with tab1:
        st.markdown('<div class="form-container">', unsafe_allow_html=True)
        
        # Toggle Switch for Login/Signup
        st.markdown("""
        <style>
            /* Toggle container */
            .toggle-container {
                display: flex;
                justify-content: center;
                margin-bottom: 2rem;
                gap: 0;
            }
            
            /* Hide radio buttons */
            div[data-testid="stRadio"] > div {
                flex-direction: row !important;
                justify-content: center !important;
                gap: 0 !important;
                background: rgba(30, 41, 59, 0.5) !important;
                padding: 0.4rem !important;
                border-radius: 12px !important;
            }
            
            div[data-testid="stRadio"] label {
                background: transparent !important;
                padding: 0.7rem 2.5rem !important;
                border-radius: 8px !important;
                cursor: pointer !important;
                transition: all 0.3s ease !important;
                color: #94a3b8 !important;
                font-weight: 500 !important;
                font-size: 1rem !important;
                margin: 0 !important;
            }
            
            div[data-testid="stRadio"] label:hover {
                color: #cbd5e1 !important;
            }
            
            div[data-testid="stRadio"] label[data-checked="true"] {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
                color: white !important;
                box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4) !important;
            }
        </style>
        """, unsafe_allow_html=True)
        
        # Toggle between Login and Signup
        user_mode = st.radio(
            "Choose action",
            ["🔑 Login", "📝 Sign Up"],
            horizontal=True,
            key="user_mode_toggle",
            label_visibility="collapsed"
        )
        
        # Show Login Form
        if user_mode == "🔑 Login":
            st.markdown('<p class="form-title">Welcome Back!</p>', unsafe_allow_html=True)
            st.markdown('<p class="footer-text" style="margin-top: 0; margin-bottom: 1.5rem;">Login with your Office ID and password</p>', unsafe_allow_html=True)
            
            login_office_id = st.text_input(
                "Office ID *", 
                key="login_office_id", 
                placeholder="e.g. EMP001"
            )
            
            login_password = st.text_input(
                "Password *", 
                type="password", 
                key="login_password", 
                placeholder="Enter your password"
            )
            
            if st.button("🚀 Login", key="user_login_btn", type="primary"):
                if not login_office_id or not login_password:
                    st.error("⚠️ Please enter Office ID and Password")
                else:
                    try:
                        user = verify_user(login_office_id, login_password)
                        if user:
                            st.success(f"✅ Welcome back, {user['name']}!")
                            st.session_state['username'] = user['username']
                            st.session_state['emp_id'] = user['emp_id']
                            st.session_state['name'] = user['name']
                            st.session_state['role'] = 'user'
                            time.sleep(0.5)
                            st.rerun()
                        else:
                            st.error("❌ Invalid credentials. Please check your Office ID and password.")
                    except Exception as e:
                        st.error(f"❌ Login error: {str(e)}")
            
            st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
            st.markdown('<p class="footer-text">💡 Don\'t have an account? Toggle to <strong>Sign Up</strong> above</p>', unsafe_allow_html=True)
        
        # Show Signup Form
        else:
            st.markdown('<p class="form-title">Create Your Account</p>', unsafe_allow_html=True)
            st.markdown('<p class="footer-text" style="margin-top: 0; margin-bottom: 1.5rem;">Fill in your details to get started</p>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                signup_office_id = st.text_input(
                    "Office ID *", 
                    key="signup_office_id", 
                    placeholder="e.g. EMP001"
                )
                
                signup_email = st.text_input(
                    "Email", 
                    key="signup_email", 
                    placeholder="your.email@company.com"
                )
                
                signup_role = st.text_input(
                    "Role", 
                    key="signup_role", 
                    placeholder="e.g. Developer"
                )
            
            with col2:
                signup_name = st.text_input(
                    "Full Name *", 
                    key="signup_name", 
                    placeholder="Your full name"
                )
                
                signup_department = st.text_input(
                    "Department", 
                    key="signup_department", 
                    placeholder="e.g. Engineering"
                )
            
            signup_password = st.text_input(
                "Password *", 
                type="password", 
                key="signup_password", 
                placeholder="Create password (min 6 chars)"
            )
            
            signup_confirm = st.text_input(
                "Confirm Password *", 
                type="password", 
                key="signup_confirm", 
                placeholder="Re-enter password"
            )
            
            if st.button("✨ Create Account", key="user_signup_btn", type="primary"):
                if not signup_office_id or not signup_name or not signup_password:
                    st.error("⚠️ Please fill in Office ID, Full Name, and Password")
                elif signup_password != signup_confirm:
                    st.error("❌ Passwords do not match!")
                elif len(signup_password) < 6:
                    st.error("❌ Password must be at least 6 characters long")
                else:
                    try:
                        new_user = add_user(
                            username=signup_office_id,
                            password=signup_password,
                            name=signup_name,
                            emp_id=signup_office_id,
                            email=signup_email or None,
                            department=signup_department or None,
                            role=signup_role or None
                        )
                        st.success(f"✅ Account created successfully for {signup_name}!")
                        st.info("🎉 Toggle to Login above to access your account")
                        time.sleep(2)
                        st.rerun()
                    except Exception as e:
                        if "already exists" in str(e).lower():
                            st.error("❌ This Office ID is already registered. Please use a different ID.")
                        else:
                            st.error(f"❌ Error creating account: {str(e)}")
            
            st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
            st.markdown('<p class="footer-text">💡 Already have an account? Toggle to <strong>Login</strong> above</p>', unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # TAB 2: ADMIN ACCESS (Login Only)
    with tab2:
        st.markdown('<div class="form-container">', unsafe_allow_html=True)
        st.markdown('<p class="form-title">Admin Login</p>', unsafe_allow_html=True)
        st.markdown('<p class="footer-text" style="margin-top: 0; margin-bottom: 1.5rem;">For authorized administrators only</p>', unsafe_allow_html=True)
        
        admin_username = st.text_input(
            "Admin Username *", 
            key="admin_username", 
            placeholder="Enter your admin username"
        )
        
        admin_password = st.text_input(
            "Admin Password *", 
            type="password", 
            key="admin_password", 
            placeholder="Enter your admin password"
        )
        
        if st.button("🔐 Admin Login", key="admin_login_btn", type="primary"):
            if not admin_username or not admin_password:
                st.error("⚠️ Please enter both username and password")
            else:
                try:
                    admin = verify_admin(admin_username, admin_password)
                    if admin:
                        st.success(f"✅ Welcome, Admin {admin['name']}!")
                        st.session_state['username'] = admin['username']
                        st.session_state['name'] = admin['name']
                        st.session_state['role'] = 'admin'
                        time.sleep(0.5)
                        st.rerun()
                    else:
                        st.error("❌ Invalid admin credentials. Please try again.")
                except Exception as e:
                    st.error(f"❌ Login error: {str(e)}")
        
        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        st.markdown('<p class="footer-text">🔒 <strong>Note:</strong> Admin accounts must be created by system administrators.</p>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

def show_user_dashboard():
    """User dashboard with attendance, submit report, and view previous reports"""
    # Sidebar
    with st.sidebar:
        st.markdown(f"### 👤 Welcome, {st.session_state.get('name', 'User')}!")
        st.markdown(f"**Employee ID:** {st.session_state.get('emp_id', 'N/A')}")
        st.markdown("---")
        page = st.radio("Navigation", ["📅 Attendance", "📝 Submit Report", "📊 My Reports"], label_visibility="collapsed")
        st.markdown("---")
        if st.button("🚪 Logout", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
        st.markdown("---")
        st.caption(f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    emp_id = st.session_state.get('emp_id')
    
    if page == "📅 Attendance":
        st.title("📅 Attendance Management")
        st.markdown("---")
        
        today = datetime.now().date()
        today_str = today.isoformat()
        attendance = get_attendance_record(emp_id, today_str)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            <div class="metric-card">
                <div class="metric-label">Check-in Status</div>
            </div>
            """, unsafe_allow_html=True)
            
            if attendance and attendance.get("checkin_time"):
                st.success(f"✅ Checked in at {attendance['checkin_time']}")
                st.info(f"📍 Status: Present")
            else:
                st.warning("⚠️ Not checked in yet")
                if st.button("🕐 Check-in Now", use_container_width=True):
                    try:
                        time_str = datetime.now().strftime('%H:%M:%S')
                        checkin_attendance(emp_id, today_str, time_str)
                        st.success("✅ Checked in successfully!")
                        time.sleep(1)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")
        
        with col2:
            st.markdown("""
            <div class="metric-card">
                <div class="metric-label">Check-out Status</div>
            </div>
            """, unsafe_allow_html=True)
            
            if attendance and attendance.get("checkout_time"):
                st.info(f"✅ Checked out at {attendance['checkout_time']}")
            else:
                if attendance and attendance.get("checkin_time"):
                    if st.button("🕐 Check-out Now", use_container_width=True):
                        try:
                            time_str = datetime.now().strftime('%H:%M:%S')
                            checkout_attendance(emp_id, today_str, time_str)
                            st.success("✅ Checked out successfully!")
                            time.sleep(1)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
                else:
                    st.warning("⚠️ Please check-in first")
        
        st.markdown("---")
        st.subheader("📊 Your Attendance History")
        
        # Get user's attendance history
        user_attendance = get_user_attendance_history(emp_id, 30)
        if user_attendance:
            att_data = []
            for a in user_attendance:
                checkin_val = a.get('checkin_time', 'Not checked in')
                checkout_val = a.get('checkout_time', 'Not checked out')
                present = a.get('checkin_time') is not None
                att_data.append({
                    'Date': a['date'],
                    'Check-in': checkin_val if checkin_val else 'Not checked in',
                    'Check-out': checkout_val if checkout_val else 'Not checked out',
                    'Status': '✅ Present' if present else '❌ Absent'
                })
            att_df = pd.DataFrame(att_data)
            st.dataframe(att_df, use_container_width=True, height=400)
        else:
            st.info("No attendance records found")
    
    elif page == "📝 Submit Report":
        show_submit_report()
    
    elif page == "📊 My Reports":
        st.title("📊 My Reports")
        st.markdown("---")
        
        try:
            # Load Excel data and filter by emp_id
            df = read_excel_data()
            if df is not None and not df.empty and 'Emp Id' in df.columns:
                # Normalize types: sometimes Emp Id is numeric in Excel while session emp_id is string
                try:
                    emp_id_str = str(emp_id).strip()
                    user_df = df[df['Emp Id'].astype(str).str.strip() == emp_id_str]
                except Exception:
                    # Fallback to original comparison if anything goes wrong
                    user_df = df[df['Emp Id'] == emp_id]
                if not user_df.empty:
                    st.subheader(f"📋 Total Reports: {len(user_df)}")
                    show_data_table(user_df)
                else:
                    st.info("📋 No reports found. Start submitting your daily reports!")
            else:
                st.info("📋 No reports found. Start submitting your daily reports!")
        except Exception as e:
            st.error(f"Error loading reports: {e}")

def show_admin_dashboard():
    """Admin dashboard with attendance reports, performance dashboard, and settings"""
    # Sidebar
    with st.sidebar:
        st.markdown(f"### 🛡️ Admin Panel")
        st.markdown(f"**Welcome, {st.session_state.get('name', 'Admin')}!**")
        st.markdown("---")
        page = st.radio("Navigation", ["👥 Attendance Report","📊 Dashboard",  "⚙️ Settings", "📧 Reminders"], label_visibility="collapsed")
        st.markdown("---")
        if st.button("🚪 Logout", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
        st.markdown("---")
        st.caption(f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    if page == "📊 Dashboard":
        st.title("📊 Employee Progress Dashboard")
        st.markdown("---")
        
        # Load all data from Excel
        df = read_excel_data()
        
        if df is None or df.empty:
            st.info("📋 No data available yet. Employees need to submit reports.")
            return
        
        # Show metrics
        show_metrics(df)
        st.markdown("---")
        
        # Show filters
        filtered_df = show_filters(df)
        st.markdown("---")
        
        # Show charts
        show_charts(filtered_df)
        st.markdown("---")
        
        # Show employee performance explorer (merged from Performance page)
        show_employee_dashboard(filtered_df if filtered_df is not None and not filtered_df.empty else df)
        st.markdown("---")
        
        # Show data table
        show_data_table(filtered_df)
    
    elif page == "👥 Attendance Report":
        st.title("👥 Employee Attendance Report")
        st.markdown("---")
        
        # TODAY'S ATTENDANCE SECTION - Show current day status first
        st.markdown("### 🔴 TODAY'S ATTENDANCE")
        st.markdown(f"**Date:** {datetime.now().strftime('%A, %B %d, %Y')}")
        
        today = datetime.now().date()
        today_str = today.isoformat()
        
        # Get today's attendance records
        attendance_data = load_attendance()
        today_records = []
        for key, record in attendance_data.items():
            if record.get('date') == today_str:
                today_records.append(record)
        
        # Get all registered users
        all_users_dict = get_all_users()
        all_users = list(all_users_dict.values())
        emp_name_map = {u['emp_id']: u['name'] for u in all_users if 'emp_id' in u and 'name' in u}
        
        if today_records or all_users:
            # Create attendance status for today
            today_att_data = []
            checked_emp_ids = set()
            
            for a in today_records:
                emp_id = a.get('emp_id')
                checked_emp_ids.add(emp_id)
                work_duration = "In Progress"
                status_icon = "🟡"
                status_text = "Checked In"
                
                checkin_time = a.get('checkin_time')
                checkout_time = a.get('checkout_time')
                
                if checkin_time and checkout_time:
                    try:
                        checkin_dt = datetime.strptime(checkin_time, '%H:%M:%S')
                        checkout_dt = datetime.strptime(checkout_time, '%H:%M:%S')
                        duration = checkout_dt - checkin_dt
                        hours = duration.total_seconds() / 3600
                        work_duration = f"{hours:.2f} hrs"
                        status_icon = "🟢"
                        status_text = "Completed"
                    except:
                        work_duration = "Invalid time"
                elif checkin_time:
                    try:
                        checkin_dt = datetime.strptime(checkin_time, '%H:%M:%S')
                        now_time = datetime.now()
                        # Combine with today's date
                        checkin_full = datetime.combine(today, checkin_dt.time())
                        duration = now_time - checkin_full
                        hours = duration.total_seconds() / 3600
                        work_duration = f"{hours:.2f} hrs (ongoing)"
                    except:
                        work_duration = "In Progress"
                
                today_att_data.append({
                    'Status': status_icon,
                    'Employee ID': emp_id,
                    'Employee Name': emp_name_map.get(emp_id, 'Unknown'),
                    'Check-in': checkin_time if checkin_time else '-',
                    'Check-out': checkout_time if checkout_time else '-',
                    'Duration': work_duration,
                    'Status Text': status_text
                })
            
            # Add employees who haven't checked in
            for user in all_users:
                emp_id = user.get('emp_id')
                if emp_id and emp_id not in checked_emp_ids:
                    today_att_data.append({
                        'Status': '🔴',
                        'Employee ID': emp_id,
                        'Employee Name': user.get('name', 'Unknown'),
                        'Check-in': 'Not checked in',
                        'Check-out': '-',
                        'Duration': '-',
                        'Status Text': 'Absent'
                    })
            
            today_df = pd.DataFrame(today_att_data)
            
            # Today's metrics - Use unique employee IDs to avoid duplicate counting
            total_emp = len(all_users)
            unique_checked_in = len(checked_emp_ids)  # Use set to count unique employees
            completed = len([r for r in today_records if r.get('checkin_time') and r.get('checkout_time')])
            in_progress = unique_checked_in - completed
            absent = total_emp - unique_checked_in
            
            # Calculate attendance rate properly
            attendance_today = round((unique_checked_in / total_emp * 100) if total_emp > 0 else 0, 1)
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown("""
                <div class="metric-card" style="background: linear-gradient(135deg, #10b981 0%, #059669 100%);">
                    <div class="metric-value">{}</div>
                    <div class="metric-label">🟢 Completed</div>
                </div>
                """.format(completed), unsafe_allow_html=True)
            
            with col2:
                st.markdown("""
                <div class="metric-card" style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);">
                    <div class="metric-value">{}</div>
                    <div class="metric-label">🟡 In Progress</div>
                </div>
                """.format(in_progress), unsafe_allow_html=True)
            
            with col3:
                st.markdown("""
                <div class="metric-card" style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);">
                    <div class="metric-value">{}</div>
                    <div class="metric-label">🔴 Absent</div>
                </div>
                """.format(absent), unsafe_allow_html=True)
            
            with col4:
                st.markdown("""
                <div class="metric-card" style="background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);">
                    <div class="metric-value">{}%</div>
                    <div class="metric-label">Today's Rate</div>
                </div>
                """.format(attendance_today), unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Display today's attendance table with color coding
            def highlight_today_status(row):
                if row['Status'] == '🟢':
                    return ['background-color: rgba(16, 185, 129, 0.15)'] * len(row)
                elif row['Status'] == '🟡':
                    return ['background-color: rgba(245, 158, 11, 0.15)'] * len(row)
                elif row['Status'] == '🔴':
                    return ['background-color: rgba(239, 68, 68, 0.15)'] * len(row)
                return [''] * len(row)
            
            styled_today = today_df.style.apply(highlight_today_status, axis=1)
            st.dataframe(styled_today, use_container_width=True, height=350)
            
            # Quick export for today
            csv_today = today_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="📥 Download Today's Attendance",
                data=csv_today,
                file_name=f"attendance_today_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        else:
            st.info("No employees found in the system.")
        
        st.markdown("---")
        st.markdown("### 📅 HISTORICAL ATTENDANCE RECORDS")
        
        # Date range filter for historical data
        col_filter1, col_filter2, col_filter3 = st.columns([2, 2, 1])
        
        with col_filter1:
            start_date = st.date_input(
                "Start Date",
                value=datetime.now().date() - timedelta(days=30),
                max_value=datetime.now().date()
            )
        
        with col_filter2:
            end_date = st.date_input(
                "End Date",
                value=datetime.now().date(),
                max_value=datetime.now().date()
            )
        
        with col_filter3:
            # Employee filter
            all_emp_ids = sorted(list(set([u.get('emp_id') for u in all_users if u.get('emp_id')])))
            selected_emp = st.selectbox("Employee", ["All"] + all_emp_ids)
        
        # Get historical records
        historical_records = []
        for key, record in attendance_data.items():
            record_date_str = record.get('date')
            if record_date_str:
                try:
                    record_date = datetime.fromisoformat(record_date_str).date()
                    if start_date <= record_date <= end_date:
                        if selected_emp == "All" or record.get('emp_id') == selected_emp:
                            historical_records.append(record)
                except:
                    pass
        
        if historical_records:
            hist_data = []
            for a in historical_records:
                emp_id = a.get('emp_id')
                checkin = a.get('checkin_time', '-')
                checkout = a.get('checkout_time', '-')
                present = checkin != '-' and checkin is not None
                
                hist_data.append({
                    'Date': a.get('date'),
                    'Employee ID': emp_id,
                    'Employee Name': emp_name_map.get(emp_id, 'Unknown'),
                    'Check-in': checkin if checkin else '-',
                    'Check-out': checkout if checkout else '-',
                    'Status': '✅ Present' if present else '❌ Absent'
                })
            
            hist_df = pd.DataFrame(hist_data)
            hist_df = hist_df.sort_values('Date', ascending=False)
            
            st.dataframe(hist_df, use_container_width=True, height=400)
            
            # Export historical data
            csv_hist = hist_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="📥 Download Historical Attendance",
                data=csv_hist,
                file_name=f"attendance_history_{start_date}_{end_date}.csv",
                mime="text/csv",
                use_container_width=True
            )
        else:
            st.info("No attendance records found for the selected period.")
    
    elif page == "⚙️ Settings":
        show_settings()
    
    elif page == "📧 Reminders":
        st.title("📧 Email Reminders")
        st.markdown("---")
        
        st.info("💡 Reminder functionality is currently disabled in local mode. Configure your email settings in config.json to enable reminders.")
        
        # Check who hasn't submitted today
        df = read_excel_data()
        today = datetime.now().date()
        
        if df is not None and not df.empty:
            if 'Date' in df.columns:
                df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                today_submissions = df[df['Date'].dt.date == today]
                submitted_ids = set(today_submissions['Emp Id'].unique()) if 'Emp Id' in today_submissions.columns else set()
                
                all_users_dict = get_all_users()
                all_emp_ids = [u.get('emp_id') for u in all_users_dict.values() if u.get('emp_id')]
                
                missing_emp_ids = [emp_id for emp_id in all_emp_ids if emp_id not in submitted_ids]
                
                if missing_emp_ids:
                    st.warning(f"⚠️ {len(missing_emp_ids)} employee(s) have not submitted reports today")
                    
                    missing_data = []
                    for emp_id in missing_emp_ids:
                        user = all_users_dict.get(emp_id, {})
                        missing_data.append({
                            'Employee ID': emp_id,
                            'Name': user.get('name', 'Unknown'),
                            'Email': user.get('email', 'Not provided')
                        })
                    
                    missing_df = pd.DataFrame(missing_data)
                    st.dataframe(missing_df, use_container_width=True)
                else:
                    st.success("✅ All employees have submitted their reports today!")
        else:
            st.error("Failed to load data")
    
    elif page == "⚙️ Settings":
        show_settings()
    
    elif page == "📧 Reminders":
        st.title("📧 Reminder Management")
        st.info("""
**Reminder System Setup**

The reminder system will automatically send emails to employees who haven't submitted their daily report.

To enable automated reminders:
1. Set up reminder time and days in Settings
2. Configure employee emails
3. Run the reminder service: `python reminder_service.py`
""")
        
        # Manual reminder test
        st.subheader("🧪 Test Reminder")
        if st.button("Check Missing Reports Today"):
            with st.spinner("Checking..."):
                df = read_excel_data()
                if df is not None:
                    missing = get_missing_reporters(df, datetime.now())
                    if missing:
                        st.warning(f"📋 {len(missing)} employees haven't reported today:")
                        for emp in missing:
                            st.write(f"- {emp}")
                    else:
                        st.success("✅ All employees have submitted their reports today!")
                else:
                    st.error("Failed to load data")

def main():
    """Main application with authentication"""
    # Check if user is logged in
    if 'role' not in st.session_state:
        show_login_signup()
        return
    
    role = st.session_state['role']
    
    if role == 'user':
        show_user_dashboard()
    elif role == 'admin':
        show_admin_dashboard()

if __name__ == "__main__":
    main()